"""
Semillado (fuente de verdad): importa el Excel UNA vez a precios.db + apus.db.

Es guardado: si las bases ya tienen datos mantenidos, se niega salvo force=True.
Aplica las correcciones de código (4613→3017) a la composición antes de insertarla.
Reutiliza el parser de pestañas del histórico (antes en ingest.py).
"""
from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, Iterator, Optional

import openpyxl

from apu_tool import config
from apu_tool.datos import correcciones
from apu_tool.datos.almacen import Almacen
from apu_tool.nucleo.models import Apu, ApuComponent, Insumo
from apu_tool.nucleo.texto import normalizar


# ---------------------------------------------------------------------------
# Excepción de semillado guardado
# ---------------------------------------------------------------------------
class SeedExistente(Exception):
    """Las bases ya tienen datos mantenidos; usar force=True para sobrescribir."""


# ---------------------------------------------------------------------------
# Utilidades de parseo (movidas desde ingest.py sin cambiar su lógica)
# ---------------------------------------------------------------------------
def _num(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace("$", "").replace(" ", "")
    # Manejo de separadores de miles/decimales comunes en es-CO.
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _code(value) -> Optional[str]:
    """Normaliza un código de insumo/apu a texto, o None si no es válido."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    return s or None


def _text(value) -> str:
    return "" if value is None else str(value).strip()


def _looks_like_code(value) -> bool:
    """Un código suele ser numérico o alfanumérico corto (INT1, etc.)."""
    c = _code(value)
    if not c:
        return False
    if c.isdigit():
        return True
    # códigos tipo INT1, INT2
    return len(c) <= 8 and any(ch.isalpha() for ch in c) and any(ch.isdigit() for ch in c)


# ---------------------------------------------------------------------------
# Configuración por pestaña (movida desde ingest.py sin cambiar su lógica)
# ---------------------------------------------------------------------------
@dataclass
class InsumoSheet:
    name: str
    col_codigo: int
    col_nombre: int
    col_unidad: int
    col_precio: int
    col_fuente: int
    col_grupo: Optional[int] = None
    start_row: int = 2


# Orden importante: la primera ocurrencia de una identidad (código, nombre) gana.
# 'INSUMOS_IDU-INT' va PRIMERO: es la base autoritativa con todos los insumos.
INSUMO_SHEETS = [
    InsumoSheet("INSUMOS_IDU-INT", col_codigo=2, col_nombre=3, col_unidad=4,
                col_precio=5, col_fuente=6, col_grupo=1),
    InsumoSheet("listado_insumos_idu", col_codigo=2, col_nombre=3, col_unidad=4,
                col_precio=5, col_fuente=6, col_grupo=1),
    InsumoSheet("insumos_apus_especificos", col_codigo=2, col_nombre=3, col_unidad=4,
                col_precio=5, col_fuente=6),
    InsumoSheet("listado_apus_idu_especiales", col_codigo=2, col_nombre=3, col_unidad=4,
                col_precio=5, col_fuente=6),
]

# Pestaña APUS: ITEM/ACTIVIDAD | COD IDU | UN | INSUMO | COD | UND | RENDIMIENTO |
#               INV | PRECIO UNITARIO | COSTO TOTAL | DIURNO/NOCTURNO
APUS_SHEET = "APUS"
APUS_COLS = dict(
    actividad=0, cod_idu=1, unidad=2, insumo_nombre=3, insumo_cod=4,
    insumo_und=5, rendimiento=6, precio_unitario=8, shift=10,
)


# ---------------------------------------------------------------------------
# Lectores (movidos desde ingest.py sin cambiar su lógica)
# ---------------------------------------------------------------------------
def _read_insumos(wb, sheet: InsumoSheet) -> Iterator[Insumo]:
    if sheet.name not in wb.sheetnames:
        return
    ws = wb[sheet.name]
    for row in ws.iter_rows(min_row=sheet.start_row, values_only=True):
        if len(row) <= sheet.col_precio:
            continue
        codigo = _code(row[sheet.col_codigo])
        if not codigo or not _looks_like_code(row[sheet.col_codigo]):
            continue
        nombre = _text(row[sheet.col_nombre])
        if not nombre:
            continue
        grupo = _text(row[sheet.col_grupo]) if sheet.col_grupo is not None else ""
        fuente = _text(row[sheet.col_fuente]) if len(row) > sheet.col_fuente else ""
        yield Insumo(
            codigo=codigo, nombre=nombre,
            unidad=_text(row[sheet.col_unidad]), grupo=grupo,
            precio=_num(row[sheet.col_precio]), fuente_precio=fuente,
        )


def _read_apus(wb) -> tuple[list[Apu], list[ApuComponent]]:
    """Recorre la pestaña APUS: cada fila con COD IDU abre un APU; las siguientes
    filas (con INSUMO/COD) son sus componentes hasta el próximo encabezado."""
    if APUS_SHEET not in wb.sheetnames:
        return [], []
    ws = wb[APUS_SHEET]
    c = APUS_COLS
    apus: dict[tuple[str, str], Apu] = {}
    comps: list[ApuComponent] = []
    cur_cod: Optional[str] = None
    cur_shift: str = config.SHIFT_DIURNO

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= c["shift"]:
            row = tuple(row) + (None,) * (c["shift"] + 1 - len(row))
        cod_idu = _code(row[c["cod_idu"]])
        actividad = _text(row[c["actividad"]])

        # ¿Fila de encabezado de APU? Tiene COD IDU y nombre de actividad.
        if cod_idu and actividad and _looks_like_code(row[c["cod_idu"]]):
            shift = _text(row[c["shift"]]).upper() or config.SHIFT_DIURNO
            shift = config.SHIFT_NOCTURNO if "NOC" in shift else config.SHIFT_DIURNO
            cur_cod, cur_shift = cod_idu, shift
            apus[(cur_cod, cur_shift)] = Apu(
                codigo=cur_cod, nombre=actividad,
                unidad=_text(row[c["unidad"]]), shift=cur_shift,
            )
            continue

        # ¿Fila de componente? Pertenece al APU abierto.
        insumo_cod = _code(row[c["insumo_cod"]])
        insumo_nombre = _text(row[c["insumo_nombre"]])
        if cur_cod and insumo_nombre:
            comps.append(ApuComponent(
                apu_codigo=cur_cod, shift=cur_shift,
                insumo_codigo=insumo_cod or "", insumo_nombre=insumo_nombre,
                unidad=_text(row[c["insumo_und"]]),
                rendimiento=_num(row[c["rendimiento"]]),
                precio_unitario_hist=_num(row[c["precio_unitario"]]),
            ))
    return list(apus.values()), comps


# ---------------------------------------------------------------------------
# API pública
# ---------------------------------------------------------------------------
def seed(almacen: Optional[Almacen] = None, xlsx_path: Optional[Path] = None,
         force: bool = False) -> dict:
    config.ensure_dirs()
    alm = almacen or Almacen()
    # Sondeo defensivo: la base puede no existir, estar vacía o traer un esquema
    # viejo. NO llamamos init_schema() aquí — crearía índices nuevos sobre tablas
    # viejas y reventaría. reset() (más abajo) es la autoridad del esquema.
    try:
        c = alm.counts()
    except sqlite3.OperationalError:
        c = {}
    if (c.get("apus", 0) or c.get("insumos", 0)) and not force:
        raise SeedExistente(
            "precios.db/apus.db ya tienen datos. Usa --force para re-semillar "
            "(¡borra correcciones mantenidas!).")

    xlsx_path = Path(xlsx_path) if xlsx_path else config.detect_source_xlsx()
    if not xlsx_path or not xlsx_path.exists():
        raise FileNotFoundError(
            "No se encontró el Excel histórico. Pásalo con --xlsx <ruta> "
            "o define la variable APU_SOURCE_XLSX.")

    alm.reset()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        insumos: dict[tuple[str, str], Insumo] = {}
        for sheet in INSUMO_SHEETS:
            for ins in _read_insumos(wb, sheet):
                insumos.setdefault((ins.codigo, normalizar(ins.nombre)), ins)
        alm.precios.insert_insumos(insumos.values())

        apus, comps = _read_apus(wb)
        alm.apus.insert_apus(apus)
        alm.apus.insert_components(correcciones.aplicar(comps))   # ← corrección aquí
    finally:
        wb.close()

    counts = alm.counts()
    alm.precios.set_meta("fuente", xlsx_path.name)
    alm.precios.set_meta("fecha_seed", date.today().isoformat())
    alm.apus.set_meta("fuente", xlsx_path.name)
    alm.apus.set_meta("fecha_seed", date.today().isoformat())
    return counts
