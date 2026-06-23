"""
Lectura de la lista de licitación (entrada) y generación de un ejemplo.

La lista trae: ítem, descripción, unidad, cantidad y precio contractual unitario.
El turno (DIURNO/NOCTURNO) puede venir por columna o como parámetro global.

El lector es tolerante con los nombres de columna: mapea encabezados por palabras
clave (es-CO), así sirve para licitaciones con formatos distintos (escalabilidad).
"""
from __future__ import annotations

import csv
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl

from apu_tool import config
from apu_tool.nucleo.models import LicitacionItem


def _norm(s: str) -> str:
    s = "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                if unicodedata.category(c) != "Mn")
    return s.strip().lower()


# Palabras clave -> campo lógico.
_HEADER_KEYS = {
    "item": ["item", "no", "numero", "codigo", "cod", "id"],
    "descripcion": ["descripcion", "actividad", "concepto", "detalle", "nombre"],
    "unidad": ["unidad", "und", "um", "u m"],
    "cantidad": ["cantidad", "cant", "qty"],
    "precio_contractual": ["precio", "valor unitario", "vr unitario", "p unitario",
                           "precio unitario", "unitario", "contractual"],
    "shift": ["turno", "jornada", "diurno", "nocturno", "shift"],
}


def _map_headers(headers: list[str]) -> dict[str, int]:
    norm_headers = [_norm(h) for h in headers]
    mapping: dict[str, int] = {}
    for field, keys in _HEADER_KEYS.items():
        for idx, h in enumerate(norm_headers):
            if any(k == h or k in h for k in keys):
                mapping[field] = idx
                break
    return mapping


def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _shift_value(raw: str, default: str) -> str:
    n = _norm(raw)
    if "noc" in n:
        return config.SHIFT_NOCTURNO
    if "diur" in n:
        return config.SHIFT_DIURNO
    return default


def _rows_from_xlsx(path: Path) -> list[list]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_from_csv(path: Path) -> list[list]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]


def read_licitacion(path: Path | str, default_shift: str = config.SHIFT_DIURNO
                    ) -> list[LicitacionItem]:
    path = Path(path)
    rows = _rows_from_xlsx(path) if path.suffix.lower() in (".xlsx", ".xlsm") \
        else _rows_from_csv(path)
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return []

    headers = [str(c) if c is not None else "" for c in rows[0]]
    mapping = _map_headers(headers)
    if "descripcion" not in mapping:
        raise ValueError(
            "No se encontró la columna de descripción/actividad en la lista. "
            f"Encabezados detectados: {headers}"
        )

    items: list[LicitacionItem] = []
    for i, row in enumerate(rows[1:], start=1):
        def get(field, default=""):
            idx = mapping.get(field)
            if idx is None or idx >= len(row):
                return default
            return row[idx]

        desc = str(get("descripcion") or "").strip()
        if not desc:
            continue
        shift = _shift_value(str(get("shift", "")), default_shift)
        items.append(LicitacionItem(
            item=str(get("item", i) or i).strip(),
            descripcion=desc,
            unidad=str(get("unidad", "") or "").strip(),
            cantidad=_to_float(get("cantidad", 1)) or 1.0,
            precio_contractual=_to_float(get("precio_contractual", 0)),
            shift=shift,
        ))
    return items


def write_sample_licitacion(path: Path | str, items: list[LicitacionItem]) -> Path:
    """Escribe un Excel de ejemplo con la estructura de entrada esperada."""
    path = Path(path)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LICITACION"
    ws.append(["ITEM", "DESCRIPCION", "UNIDAD", "CANTIDAD",
               "PRECIO CONTRACTUAL", "TURNO"])
    for it in items:
        ws.append([it.item, it.descripcion, it.unidad, it.cantidad,
                   it.precio_contractual, it.shift])
    wb.save(path)
    return path
