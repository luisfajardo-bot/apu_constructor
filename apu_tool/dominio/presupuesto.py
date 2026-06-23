"""
Lectura del presupuesto oficial por capítulos (hoja FOR 1-PPTO OFICIAL).

El presupuesto está organizado jerárquicamente:
    Capítulo (con número)  ->  TURNO DIURNO/NOCTURNO  ->  subgrupo  ->  ítems.

Se recorre de arriba abajo llevando el estado (capítulo, turno) vigente; cada ítem
hereda ambos. El precio contractual es el valor unitario BÁSICO (sin AIU), columna [9].
A diferencia de la licitación plana, cada ítem trae su código IDU (columna [2]), que
permite armar el APU por código directo.
"""
from __future__ import annotations

import unicodedata
from pathlib import Path

import openpyxl

from apu_tool import config
from apu_tool.nucleo.models import LicitacionItem

# Índices de columna (0-idx) en la hoja FOR 1-PPTO OFICIAL.
COL_CODIGO = 2
COL_ITEMPAGO = 3
COL_DESC = 6
COL_UND = 7
COL_CANT = 8
COL_PRECIO = 9   # valor unitario BÁSICO (sin AIU)

HOJA_DEFECTO = "FOR 1-PPTO OFICIAL"


def _norm(s) -> str:
    s = "".join(c for c in unicodedata.normalize("NFD", str(s or ""))
                if unicodedata.category(c) != "Mn")
    return s.strip().lower()


def _to_float(v) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _code(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _es_codigo_item(v) -> bool:
    """Un código de ítem del presupuesto es numérico (3009) o alfanumérico corto."""
    c = _code(v)
    if not c:
        return False
    if c.isdigit():
        return True
    return len(c) <= 8 and any(ch.isalpha() for ch in c) and any(ch.isdigit() for ch in c)


def _es_numero_capitulo(v) -> bool:
    """Capítulo: la columna de ítem de pago trae un entero (7), no un 7.101."""
    if isinstance(v, int):
        return True
    if isinstance(v, float):
        return v.is_integer()
    s = str(v or "").strip()
    return s.isdigit()


def _get(row: list, idx: int):
    return row[idx] if idx < len(row) else None


def read_presupuesto(path: Path | str, hoja: str = HOJA_DEFECTO,
                     default_shift: str = config.SHIFT_DIURNO) -> list[LicitacionItem]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if hoja not in wb.sheetnames:
            raise ValueError(
                f"No se encontró la hoja '{hoja}'. Hojas: {wb.sheetnames}")
        ws = wb[hoja]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    capitulo = ""
    turno = default_shift
    items: list[LicitacionItem] = []

    for row in rows:
        codigo = _code(_get(row, COL_CODIGO))
        cantidad = _to_float(_get(row, COL_CANT))
        desc = str(_get(row, COL_DESC) or "").strip()

        # Ítem: tiene código válido y cantidad > 0.
        if cantidad > 0 and _es_codigo_item(_get(row, COL_CODIGO)):
            items.append(LicitacionItem(
                item=_code(_get(row, COL_ITEMPAGO)) or codigo,
                descripcion=desc,
                unidad=str(_get(row, COL_UND) or "").strip(),
                cantidad=cantidad,
                precio_contractual=_to_float(_get(row, COL_PRECIO)),
                shift=turno,
                categoria=capitulo,
                codigo_sugerido=codigo,
            ))
            continue

        # Encabezado: hay descripción y NO hay código de ítem.
        if desc and not codigo:
            n = _norm(desc)
            if "turno" in n:
                turno = (config.SHIFT_NOCTURNO if "noc" in n else config.SHIFT_DIURNO)
            elif _es_numero_capitulo(_get(row, COL_ITEMPAGO)):
                num = _code(_get(row, COL_ITEMPAGO))
                capitulo = f"{num} · {desc}" if num else desc
            # otros encabezados (subgrupos) no cambian capítulo ni turno.
    return items
