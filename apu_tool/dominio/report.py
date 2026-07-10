"""
Generación del cuadro resumen (salida en Excel).

Contiene tres hojas:
  - RESUMEN   : una fila por ítem con contractual vs costo y margen, + totales.
  - DESGLOSE  : la composición costeada de cada APU (insumos, rendimiento, precio).
  - ALERTAS   : ítems que requieren revisión o armado manual.

El cuadro de resumen es el entregable principal pedido por el usuario.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apu_tool.dominio.alertas import alertas_costeo, filas_alertadas, motivo_alerta
from apu_tool.nucleo.models import AssembledApu, MatchStatus

_MONEY = '#,##0'
_PCT = '0.0%'
_REND = '#,##0.0000'

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")
_BAD_FILL = PatternFill("solid", fgColor="FCE4E4")     # margen negativo
_WARN_FILL = PatternFill("solid", fgColor="FFF2CC")    # revisar
_ALERT_FILL = PatternFill("solid", fgColor="F8CBAD")   # naranja: alerta de costeo ($0 / cruce dudoso)
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STATUS_LABEL = {
    MatchStatus.AUTO: "Automático",
    MatchStatus.REVIEW: "Revisar",
    MatchStatus.NEW: "Manual",
    MatchStatus.CONFIRMED: "Confirmado",
    MatchStatus.REJECTED: "Rechazado",
}


def _style_header(ws, row: int, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border = _BORDER


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _build_resumen(ws, apus: list[AssembledApu]) -> None:
    headers = ["Ítem", "Descripción", "Und", "Cantidad",
               "P. Contractual", "Costo Unit.", "Margen Unit.", "Margen %",
               "Total Contractual", "Total Costo", "Margen Total",
               "Estado", "Confianza", "APU base"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for a in apus:
        ws.append([
            a.item.item, a.item.descripcion, a.unidad, a.item.cantidad,
            a.item.precio_contractual, a.costo_unitario, a.margen_unitario,
            a.margen_pct, a.contractual_total, a.costo_total, a.margen_total,
            _STATUS_LABEL.get(a.status, a.status), round(a.confianza, 2),
            a.apu_codigo or "",
        ])
        r = ws.max_row
        for col in (4,):
            ws.cell(row=r, column=col).number_format = _REND
        for col in (5, 6, 7, 9, 10, 11):
            ws.cell(row=r, column=col).number_format = _MONEY
        ws.cell(row=r, column=8).number_format = _PCT
        # Resaltado por prioridad: alerta de costeo > margen negativo > revisar.
        if alertas_costeo(a):
            fill = _ALERT_FILL
        elif a.margen_total < 0:
            fill = _BAD_FILL
        elif a.status in (MatchStatus.REVIEW, MatchStatus.NEW):
            fill = _WARN_FILL
        else:
            fill = None
        if fill is not None:
            for col in range(1, len(headers) + 1):
                ws.cell(row=r, column=col).fill = fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).border = _BORDER

    # Fila de totales.
    total_contractual = sum(a.contractual_total for a in apus)
    total_costo = sum(a.costo_total for a in apus)
    total_margen = total_contractual - total_costo
    ws.append([])
    ws.append(["", "TOTALES", "", "", "", "", "", "",
               total_contractual, total_costo, total_margen, "", "", ""])
    r = ws.max_row
    for col in (9, 10, 11):
        c = ws.cell(row=r, column=col)
        c.number_format = _MONEY
        c.font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).fill = _TOTAL_FILL
    ws.cell(row=r, column=2).font = Font(bold=True)

    # Margen % global.
    ws.append(["", "MARGEN % GLOBAL", "", "", "", "", "", "",
               "", "", (total_margen / total_contractual) if total_contractual else 0])
    ws.cell(row=ws.max_row, column=11).number_format = _PCT
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    _autosize(ws, {1: 8, 2: 50, 3: 6, 4: 12, 5: 16, 6: 14, 7: 14, 8: 10,
                   9: 18, 10: 16, 11: 16, 12: 12, 13: 10, 14: 10})


def _build_desglose(ws, apus: list[AssembledApu]) -> None:
    headers = ["Ítem", "APU", "Actividad", "Insumo Cód", "Insumo",
               "Und", "Rendimiento", "Precio Unit.", "Fuente precio", "Costo", "Cruce"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    for a in apus:
        if not a.componentes:
            ws.append([a.item.item, a.apu_codigo or "", a.apu_nombre,
                       "", "(sin composición — armar manual)", "", "", "", "", "", ""])
            continue
        for c in a.componentes:
            ws.append([a.item.item, a.apu_codigo or "", a.apu_nombre,
                       c.insumo_codigo, c.insumo_nombre, c.unidad,
                       c.rendimiento, c.precio_unitario, c.fuente_precio, c.costo,
                       c.calidad_cruce])
            r = ws.max_row
            ws.cell(row=r, column=7).number_format = _REND
            ws.cell(row=r, column=8).number_format = _MONEY
            ws.cell(row=r, column=10).number_format = _MONEY
    _autosize(ws, {1: 8, 2: 8, 3: 40, 4: 10, 5: 40, 6: 8, 7: 14, 8: 14,
                   9: 18, 10: 14, 11: 12})


def _build_alertas(ws, apus: list[AssembledApu]) -> None:
    headers = ["Ítem", "Descripción", "Estado", "Confianza",
               "APU propuesto", "Justificación / motivo"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    flagged = filas_alertadas(apus)
    for a, ac in flagged:
        motivo = motivo_alerta(a, ac)
        ws.append([a.item.item, a.item.descripcion,
                   _STATUS_LABEL.get(a.status, a.status), round(a.confianza, 2),
                   a.apu_codigo or "", motivo])
        ws.cell(row=ws.max_row, column=4).number_format = '0.00'
    if not flagged:
        ws.append(["", "Sin alertas: todos los ítems se armaron con coincidencia clara.",
                   "", "", "", ""])
    _autosize(ws, {1: 8, 2: 45, 3: 12, 4: 10, 5: 14, 6: 60})


def write_report(apus: list[AssembledApu], path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    _build_resumen(wb.active, apus)
    wb.active.title = "RESUMEN"
    _build_desglose(wb.create_sheet("DESGLOSE"), apus)
    _build_alertas(wb.create_sheet("ALERTAS"), apus)
    # Metadatos.
    meta = wb.create_sheet("INFO")
    meta.append(["Generado", date.today().isoformat()])
    meta.append(["Ítems", len(apus)])
    meta.append(["Nota", "Los precios de costo NO fueron vistos por la IA. "
                         "La IA solo decidió la estructura de los APUs."])
    wb.save(path)
    return path
