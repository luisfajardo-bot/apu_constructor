"""
Cuadro resumen agrupado por capítulos del presupuesto.

Hojas:
  - RESUMEN POR CAPÍTULO : un renglón por capítulo (contractual vs costo, margen) + gran total.
  - DETALLE              : ítems agrupados bajo cada capítulo, con subtotal por capítulo.
  - APUS                 : cada APU como bloque apilado (estilo de la pestaña APUS original).
  - ALERTAS              : ítems que requieren revisión o armado manual.
  - INFO                 : metadatos + nota de privacidad (la IA no vio dinero).

Reúsa los estilos de report.py para no duplicar formato.
"""
from __future__ import annotations

from datetime import date
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

from apu_tool.nucleo.models import AssembledApu, MatchStatus
from apu_tool.dominio.report import (_MONEY, _PCT, _REND, _STATUS_LABEL, _TOTAL_FILL,
                                     _WARN_FILL, _autosize, _style_header)


def agrupar_por_capitulo(apus: list[AssembledApu]) -> dict[str, list[AssembledApu]]:
    """Agrupa por capítulo preservando el orden de aparición."""
    grupos: dict[str, list[AssembledApu]] = {}
    for a in apus:
        cap = a.item.categoria or "(sin capítulo)"
        grupos.setdefault(cap, []).append(a)
    return grupos


def _build_resumen_capitulo(ws, grupos: dict[str, list[AssembledApu]]) -> None:
    headers = ["Capítulo", "Total Contractual", "Total Costo",
               "Margen Total", "Margen %"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    g_contractual = g_costo = 0.0
    for cap, apus in grupos.items():
        c = sum(a.contractual_total for a in apus)
        k = sum(a.costo_total for a in apus)
        m = c - k
        ws.append([cap, c, k, m, (m / c) if c else 0.0])
        r = ws.max_row
        for col in (2, 3, 4):
            ws.cell(row=r, column=col).number_format = _MONEY
        ws.cell(row=r, column=5).number_format = _PCT
        g_contractual += c
        g_costo += k

    g_margen = g_contractual - g_costo
    ws.append(["GRAN TOTAL", g_contractual, g_costo, g_margen,
               (g_margen / g_contractual) if g_contractual else 0.0])
    r = ws.max_row
    for col in (2, 3, 4):
        cell = ws.cell(row=r, column=col)
        cell.number_format = _MONEY
        cell.font = Font(bold=True)
    ws.cell(row=r, column=5).number_format = _PCT
    for col in range(1, len(headers) + 1):
        ws.cell(row=r, column=col).fill = _TOTAL_FILL
    ws.cell(row=r, column=1).font = Font(bold=True)
    _autosize(ws, {1: 46, 2: 18, 3: 16, 4: 16, 5: 10})


def _build_detalle(ws, grupos: dict[str, list[AssembledApu]]) -> None:
    headers = ["Ítem", "Descripción", "Und", "Cantidad", "P. Contractual",
               "Costo Unit.", "Margen Unit.", "Margen %", "Total Contractual",
               "Total Costo", "Margen Total", "Estado"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    for cap, apus in grupos.items():
        ws.append([cap])
        cr = ws.max_row
        ws.cell(row=cr, column=1).font = Font(bold=True)
        for col in range(1, len(headers) + 1):
            ws.cell(row=cr, column=col).fill = _TOTAL_FILL

        for a in apus:
            ws.append([
                a.item.item, a.item.descripcion, a.unidad, a.item.cantidad,
                a.item.precio_contractual, a.costo_unitario, a.margen_unitario,
                a.margen_pct, a.contractual_total, a.costo_total, a.margen_total,
                _STATUS_LABEL.get(a.status, a.status),
            ])
            r = ws.max_row
            ws.cell(row=r, column=4).number_format = _REND
            for col in (5, 6, 7, 9, 10, 11):
                ws.cell(row=r, column=col).number_format = _MONEY
            ws.cell(row=r, column=8).number_format = _PCT
            if a.status in (MatchStatus.REVIEW, MatchStatus.NEW):
                for col in range(1, len(headers) + 1):
                    ws.cell(row=r, column=col).fill = _WARN_FILL

        sc = sum(a.contractual_total for a in apus)
        sk = sum(a.costo_total for a in apus)
        ws.append(["", f"Subtotal {cap}", "", "", "", "", "", "",
                   sc, sk, sc - sk, ""])
        r = ws.max_row
        for col in (9, 10, 11):
            cell = ws.cell(row=r, column=col)
            cell.number_format = _MONEY
            cell.font = Font(bold=True)
        ws.cell(row=r, column=2).font = Font(bold=True)
    _autosize(ws, {1: 9, 2: 46, 3: 6, 4: 12, 5: 16, 6: 14, 7: 14, 8: 10,
                   9: 18, 10: 16, 11: 16, 12: 12})


def _build_apus(ws, apus: list[AssembledApu]) -> None:
    """Cada APU como bloque apilado: título + insumos + costo unitario."""
    from openpyxl.styles import PatternFill
    sub = ["Insumo Cód", "Insumo", "Und", "Rendimiento", "Precio Unit.",
           "Fuente", "Costo"]
    for a in apus:
        titulo = f"{a.item.item}   {a.apu_nombre}   ({a.unidad})"
        ws.append([titulo])
        tr = ws.max_row
        ws.cell(row=tr, column=1).font = Font(bold=True, color="FFFFFF")
        for col in range(1, len(sub) + 1):
            ws.cell(row=tr, column=col).fill = PatternFill("solid", fgColor="1F4E78")

        ws.append(sub)
        _style_header(ws, ws.max_row, len(sub))

        if not a.componentes:
            ws.append(["", "(sin composición — armar manual)", "", "", "", "", ""])
        for c in a.componentes:
            ws.append([c.insumo_codigo, c.insumo_nombre, c.unidad, c.rendimiento,
                       c.precio_unitario, c.fuente_precio, c.costo])
            r = ws.max_row
            ws.cell(row=r, column=4).number_format = _REND
            ws.cell(row=r, column=5).number_format = _MONEY
            ws.cell(row=r, column=7).number_format = _MONEY

        ws.append(["", "COSTO UNITARIO APU", "", "", "", "", a.costo_unitario])
        r = ws.max_row
        ws.cell(row=r, column=2).font = Font(bold=True)
        ws.cell(row=r, column=7).number_format = _MONEY
        ws.cell(row=r, column=7).font = Font(bold=True)
        ws.append([])  # línea en blanco entre APUs
    _autosize(ws, {1: 12, 2: 46, 3: 8, 4: 14, 5: 14, 6: 18, 7: 14})


def _build_alertas(ws, apus: list[AssembledApu]) -> None:
    headers = ["Ítem", "Descripción", "Capítulo", "Estado", "Confianza",
               "APU propuesto", "Justificación / motivo"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"
    flagged = [a for a in apus if a.status in (MatchStatus.REVIEW, MatchStatus.NEW)]
    for a in flagged:
        ws.append([a.item.item, a.item.descripcion, a.item.categoria,
                   _STATUS_LABEL.get(a.status, a.status), round(a.confianza, 2),
                   a.apu_codigo or "", a.explicacion])
        ws.cell(row=ws.max_row, column=5).number_format = '0.00'
    if not flagged:
        ws.append(["", "Sin alertas: todos los ítems se armaron con coincidencia clara."])
    _autosize(ws, {1: 9, 2: 45, 3: 30, 4: 12, 5: 10, 6: 14, 7: 50})


def write_report_categorizado(apus: list[AssembledApu], path: Path | str) -> Path:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    grupos = agrupar_por_capitulo(apus)

    wb = openpyxl.Workbook()
    _build_resumen_capitulo(wb.active, grupos)
    wb.active.title = "RESUMEN POR CAPÍTULO"
    _build_detalle(wb.create_sheet("DETALLE"), grupos)
    _build_apus(wb.create_sheet("APUS"), apus)
    _build_alertas(wb.create_sheet("ALERTAS"), apus)

    info = wb.create_sheet("INFO")
    info.append(["Generado", date.today().isoformat()])
    info.append(["Ítems", len(apus)])
    info.append(["Capítulos", len(grupos)])
    info.append(["Precio contractual", "Valor unitario BÁSICO (sin AIU)"])
    info.append(["Nota", "Los precios de costo NO fueron vistos por la IA. "
                         "La IA solo decidió la estructura de los APUs."])
    wb.save(path)
    return path
