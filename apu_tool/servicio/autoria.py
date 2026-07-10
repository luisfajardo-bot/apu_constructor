"""
Lógica de servicio para AGREGAR a la base: insumos y APUs nuevos.

Altas de catálogo —individual o por Excel—, separadas de la edición de precios
(`insumos.py`) y de las corridas. Ve dinero (es catálogo del equipo), pero NUNCA
abre un camino hacia la IA (Invariante #1: la frontera de privacidad).

El import de APUs reutiliza el parser de la hoja `APUS` del seed, así que el
formato es idéntico al del histórico que el usuario ya maneja. No aplica las
correcciones de código del seed (esas son del histórico original).
"""
from __future__ import annotations

import csv
import io
from dataclasses import replace

import openpyxl

from apu_tool import config
from apu_tool.datos.almacen import Almacen
from apu_tool.datos.seed import _read_apus
from apu_tool.nucleo.models import Apu, ApuComponent, Insumo
from apu_tool.nucleo.texto import normalizar
from apu_tool.servicio.auditoria import nuevo_lote, registrar_auditoria
from apu_tool.servicio.insumos import _insumo_out, _norm_h, _to_float, MSG_PRECIO_POSITIVO
from apu_tool.servicio.subapus import (
    mapa_codigos_apu, nombres_apu, detectar_subapus_lote, marcar_comps_subapu,
)


# ----------------------------------------------------------------- individual
def crear_insumo(alm: Almacen, datos: dict, actor=None) -> dict:
    codigo = str(datos.get("codigo", "") or "").strip()
    nombre = str(datos.get("nombre", "") or "").strip()
    if not codigo or not nombre:
        raise ValueError("Código y nombre son obligatorios.")
    precio = _to_float(datos.get("precio"))
    if precio <= 0:
        raise ValueError(MSG_PRECIO_POSITIVO)
    ins = Insumo(codigo=codigo, nombre=nombre,
                 unidad=str(datos.get("unidad", "") or ""),
                 grupo=str(datos.get("grupo", "") or ""),
                 precio=precio, fuente_precio=str(datos.get("fuente", "") or ""))
    with alm.transaccion("precios") as conn:
        iid = alm.precios.crear_insumo(ins, conn=conn,
                                       creado_por=(actor.user_id if actor else None))
        registrar_auditoria(
            alm, conn, actor, "insumo.crear", "insumo", iid, antes=None,
            despues={"codigo": ins.codigo, "nombre": ins.nombre, "unidad": ins.unidad,
                     "grupo": ins.grupo, "precio": ins.precio, "fuente": ins.fuente_precio},
            contexto={"origen": "individual"})
    return _insumo_out(alm.precios.get_insumo_por_id(iid))


def _componentes_de(alm: Almacen, comp_dicts: list[dict], shift: str,
                    previos: dict | None = None) -> list[ApuComponent]:
    """Arma los ApuComponent a partir de los dicts del contrato HTTP.

    `previos` (opcional): marcas existentes por código -> (tipo, ref_shift), para
    preservarlas al editar cuando el componente entrante no trae `tipo` explícito
    (invariante de FIX 1: editar un APU no debe borrar las marcas de sub-APU)."""
    previos = previos or {}
    comps: list[ApuComponent] = []
    for c in comp_dicts:
        cod = str(c.get("insumo_codigo", "") or "").strip()
        rend = _to_float(c.get("rendimiento"))
        if not cod:
            raise ValueError("Cada componente necesita un código de insumo.")
        if rend <= 0:
            raise ValueError(f"El rendimiento del insumo {cod} debe ser mayor que 0.")
        # Resuelve nombre/unidad desde la base si el insumo existe (respaldo embebido);
        # si no existe, se guarda lo que venga (enlace blando -> cruce huérfano al costear).
        cands = alm.precios.get_candidatos(cod)
        if cands:
            nombre, unidad = cands[0].nombre, cands[0].unidad
        else:
            nombre = str(c.get("insumo_nombre", "") or "")
            unidad = str(c.get("unidad", "") or "")
        tipo_in = c.get("tipo")
        if tipo_in:
            tipo, ref_shift = str(tipo_in), str(c.get("ref_shift", "") or "")
        elif cod in previos:
            tipo, ref_shift = previos[cod]
        else:
            tipo, ref_shift = "insumo", ""
        comps.append(ApuComponent(
            apu_codigo="", shift=shift, insumo_codigo=cod, insumo_nombre=nombre,
            unidad=unidad, rendimiento=rend, precio_unitario_hist=0.0,
            tipo=tipo, ref_shift=ref_shift))
    return comps


def crear_apu(alm: Almacen, datos: dict, actor=None) -> dict:
    codigo = str(datos.get("codigo", "") or "").strip()
    nombre = str(datos.get("nombre", "") or "").strip()
    turno = str(datos.get("turno", "") or "").strip().upper()
    if not codigo or not nombre:
        raise ValueError("Código y nombre son obligatorios.")
    if turno not in (config.SHIFT_DIURNO, config.SHIFT_NOCTURNO):
        raise ValueError("El turno debe ser DIURNO o NOCTURNO.")
    comps = _componentes_de(alm, datos.get("componentes", []) or [], turno)
    apu = Apu(codigo=codigo, nombre=nombre, unidad=str(datos.get("unidad", "") or ""),
              shift=turno, grupo=str(datos.get("grupo", "") or ""))
    with alm.transaccion("apus") as conn:
        alm.apus.crear_apu(apu, comps, conn=conn)
        registrar_auditoria(
            alm, conn, actor, "apu.crear", "apu", codigo, antes=None,
            despues={"codigo": codigo, "turno": turno, "nombre": nombre,
                     "unidad": apu.unidad, "grupo": apu.grupo, "n_componentes": len(comps)},
            contexto={"origen": "individual"})
    return {"codigo": codigo, "turno": turno, "nombre": nombre,
            "unidad": apu.unidad, "grupo": apu.grupo, "n_componentes": len(comps)}


def editar_apu(alm: Almacen, codigo: str, shift: str, datos: dict, actor=None) -> dict | None:
    """Edita cabecera + composición de un APU existente. Identidad (codigo, turno) fija.
    Devuelve None si no existe (endpoint -> 404); ValueError en validación (-> 400)."""
    codigo = str(codigo or "").strip()
    shift = str(shift or "").strip().upper()
    previo = alm.apus.get_apu(codigo, shift)
    if previo is None:
        return None
    nombre = str(datos.get("nombre", "") or "").strip()
    if not nombre:
        raise ValueError("El nombre es obligatorio.")
    existentes = alm.apus.get_components(codigo, shift)
    previos: dict[str, tuple[str, str]] = {}
    for e in existentes:
        if e.insumo_codigo not in previos or e.tipo == "apu":
            previos[e.insumo_codigo] = (e.tipo, e.ref_shift)
    comps = _componentes_de(alm, datos.get("componentes", []) or [], shift, previos=previos)
    antes = {"nombre": previo.nombre, "unidad": previo.unidad, "grupo": previo.grupo,
             "n_componentes": len(existentes)}
    apu = Apu(codigo=codigo, nombre=nombre, unidad=str(datos.get("unidad", "") or ""),
              shift=shift, grupo=str(datos.get("grupo", "") or ""))
    with alm.transaccion("apus") as conn:
        alm.apus.editar_apu(apu, comps, conn=conn)
        registrar_auditoria(
            alm, conn, actor, "apu.editar", "apu", codigo, antes=antes,
            despues={"nombre": nombre, "unidad": apu.unidad, "grupo": apu.grupo,
                     "n_componentes": len(comps)},
            contexto={"origen": "individual"})
    return {"codigo": codigo, "turno": shift, "nombre": nombre,
            "unidad": apu.unidad, "grupo": apu.grupo, "n_componentes": len(comps)}


def borrar_apu(alm: Almacen, codigo: str, shift: str, actor=None) -> dict | None:
    """Borra un APU (cabecera + composición). Devuelve None si no existe (endpoint -> 404).
    Las corridas ya armadas conservan su foto; se informa cuántas lo referencian."""
    codigo = str(codigo or "").strip()
    shift = str(shift or "").strip().upper()
    previo = alm.apus.get_apu(codigo, shift)
    if previo is None:
        return None
    n_comps = len(alm.apus.get_components(codigo, shift))
    n_corridas = alm.corridas.contar_items_por_apu(codigo)
    antes = {"codigo": codigo, "turno": shift, "nombre": previo.nombre,
             "unidad": previo.unidad, "grupo": previo.grupo, "n_componentes": n_comps}
    with alm.transaccion("apus") as conn:
        alm.apus.borrar_apu(codigo, shift, conn=conn)
        registrar_auditoria(
            alm, conn, actor, "apu.borrar", "apu", codigo, antes=antes, despues=None,
            contexto={"n_corridas": n_corridas})
    return {"borrado": True, "n_corridas": n_corridas}


# ------------------------------------------------------------- import insumos
def _match_identidad(alm: Almacen, codigo: str, nombre: str):
    """Insumo con (codigo, nombre) exactos (nombre normalizado), o None."""
    nn = normalizar(nombre)
    for c in alm.precios.get_candidatos(codigo):
        if normalizar(c.nombre) == nn:
            return c
    return None


def _cambio_upsert(ins, f: dict) -> dict:
    precio_nuevo = f["precio"] if f["tiene_precio"] else ins.precio
    fuente_nueva = f["fuente"] or ins.fuente_precio
    return {"insumo_id": ins.id, "codigo": ins.codigo, "nombre": ins.nombre,
            "precio_actual": ins.precio, "precio_nuevo": precio_nuevo,
            "fuente_actual": ins.fuente_precio, "fuente_nueva": fuente_nueva}


def _filas_insumos(contenido: bytes, nombre_archivo: str) -> list[dict]:
    """Lee una tabla con columnas codigo, nombre, unidad, grupo, precio, fuente."""
    if nombre_archivo.lower().endswith((".xlsx", ".xlsm")):
        wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        rows = [list(r) for r in wb.active.iter_rows(values_only=True)]
        wb.close()
    else:
        text = contenido.decode("utf-8-sig", errors="replace")
        rows = [r for r in csv.reader(io.StringIO(text))]
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return []
    headers = [_norm_h(c) for c in rows[0]]

    def col(*keys):
        for i, h in enumerate(headers):
            if any(k == h or k in h for k in keys):
                return i
        return None

    ci = {"codigo": col("codigo", "cod", "code"),
          "nombre": col("nombre", "descripcion", "name"),
          "unidad": col("unidad", "und", "unit"),
          "grupo": col("grupo", "group"),
          "precio": col("precio", "valor", "price"),
          "fuente": col("fuente", "source")}
    if ci["codigo"] is None:
        raise ValueError("El archivo debe tener al menos una columna de código.")

    def g(r, i):
        return r[i] if (i is not None and i < len(r)) else None

    out = []
    for r in rows[1:]:
        raw_precio = g(r, ci["precio"])
        out.append({"codigo": str(g(r, ci["codigo"]) or "").strip(),
                    "nombre": str(g(r, ci["nombre"]) or "").strip(),
                    "unidad": str(g(r, ci["unidad"]) or "").strip(),
                    "grupo": str(g(r, ci["grupo"]) or "").strip(),
                    "precio": _to_float(raw_precio),
                    "tiene_precio": raw_precio not in (None, ""),
                    "fuente": str(g(r, ci["fuente"]) or "").strip()})
    return out


def preview_importar_insumos(alm: Almacen, contenido: bytes, nombre_archivo: str) -> dict:
    """Upsert por fila. Con nombre: identidad código+nombre (crea o actualiza).
    Sin nombre: actualiza precio por código (único), o marca ambigua/no encontrada."""
    crear, actualizar, ambigua, no_encontrada, invalida = [], [], [], [], []
    for f in _filas_insumos(contenido, nombre_archivo):
        cod, nom = f["codigo"], f["nombre"]
        if not cod:
            invalida.append(f)
        elif nom:
            match = _match_identidad(alm, cod, nom)
            (actualizar.append(_cambio_upsert(match, f)) if match else crear.append(f))
        else:
            cands = alm.precios.get_candidatos(cod)
            if len(cands) == 1:
                actualizar.append(_cambio_upsert(cands[0], f))
            elif len(cands) > 1:
                ambigua.append({"codigo": cod,
                                "candidatos": [{"id": c.id, "nombre": c.nombre} for c in cands]})
            else:
                no_encontrada.append({"codigo": cod})
    return {"crear": crear, "actualizar": actualizar, "ambigua": ambigua,
            "no_encontrada": no_encontrada, "invalida": invalida}


def aplicar_importar_insumos(alm: Almacen, contenido: bytes, nombre_archivo: str,
                             actor=None) -> dict:
    prev = preview_importar_insumos(alm, contenido, nombre_archivo)
    creados, actualizados, errores = 0, 0, []
    lote = nuevo_lote()
    for f in prev["crear"]:
        try:
            if f["precio"] <= 0:
                raise ValueError(MSG_PRECIO_POSITIVO)
            ins = Insumo(codigo=f["codigo"], nombre=f["nombre"], unidad=f["unidad"],
                         grupo=f["grupo"], precio=f["precio"], fuente_precio=f["fuente"])
            with alm.transaccion("precios") as conn:
                iid = alm.precios.crear_insumo(ins, conn=conn,
                                               creado_por=(actor.user_id if actor else None))
                registrar_auditoria(
                    alm, conn, actor, "insumo.crear", "insumo", iid, antes=None,
                    despues={"codigo": ins.codigo, "nombre": ins.nombre, "unidad": ins.unidad,
                             "grupo": ins.grupo, "precio": ins.precio, "fuente": ins.fuente_precio},
                    contexto={"origen": "import", "lote_id": lote, "archivo": nombre_archivo})
            creados += 1
        except ValueError as e:
            errores.append({"codigo": f["codigo"], "error": str(e)})
    for c in prev["actualizar"]:
        if c["precio_nuevo"] == c["precio_actual"] and c["fuente_nueva"] == c["fuente_actual"]:
            continue                                   # no-op: nada cambió
        try:
            if c["precio_nuevo"] <= 0:
                raise ValueError(MSG_PRECIO_POSITIVO)
            with alm.transaccion("precios") as conn:
                alm.precios.set_precio_por_id(c["insumo_id"], c["precio_nuevo"], c["fuente_nueva"],
                                              conn=conn,
                                              creado_por=(actor.user_id if actor else None))
                registrar_auditoria(
                    alm, conn, actor, "precio.editar", "insumo", c["insumo_id"],
                    antes={"precio": c["precio_actual"], "fuente": c["fuente_actual"]},
                    despues={"precio": c["precio_nuevo"], "fuente": c["fuente_nueva"]},
                    contexto={"origen": "import", "lote_id": lote, "archivo": nombre_archivo})
            actualizados += 1
        except Exception as e:
            errores.append({"codigo": c["codigo"], "error": str(e)})
    return {"creados": creados, "actualizados": actualizados, "errores": errores}


# ---------------------------------------------------------------- import APUs
def _codigo_con_turno(codigo: str, shift: str) -> str:
    """Convención de la empresa: el APU nocturno lleva el código con sufijo ' N'
    (p.ej. '3010' -> '3010 N'), igual que el histórico. La lista de licitación /
    plantilla suele traer el código PELADO y marcar lo nocturno solo en la columna
    de turno; aquí lo normalizamos para que el APU importado quede con la misma
    identidad que en la biblioteca. Idempotente: no re-sufija si ya termina en 'N'.
    (El remapeo de insumos a tarifa nocturna '4278 N' es aparte, fuera de alcance.)"""
    if shift == config.SHIFT_NOCTURNO and codigo and not codigo.rstrip().upper().endswith("N"):
        return f"{codigo.rstrip()} N"
    return codigo


def _parse_apus(contenido: bytes):
    """Parsea un Excel con hoja 'APUS' (formato del histórico) reusando el parser
    del seed. Devuelve (apus, comps_por_clave[(codigo, shift)]). Aplica la
    convención de turno al código (ver `_codigo_con_turno`)."""
    wb = openpyxl.load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    try:
        if "APUS" not in wb.sheetnames:
            raise ValueError(
                "El Excel debe tener una hoja llamada 'APUS' con el formato del histórico.")
        apus, comps = _read_apus(wb)
    finally:
        wb.close()
    apus = [replace(a, codigo=_codigo_con_turno(a.codigo, a.shift)) for a in apus]
    comps_por: dict[tuple[str, str], list[ApuComponent]] = {}
    for c in comps:
        cod = _codigo_con_turno(c.apu_codigo, c.shift)
        comp = c if cod == c.apu_codigo else replace(c, apu_codigo=cod)
        comps_por.setdefault((cod, comp.shift), []).append(comp)
    return apus, comps_por


def preview_importar_apus(alm: Almacen, contenido: bytes) -> dict:
    apus, comps_por = _parse_apus(contenido)
    crear, ya_existe, crear_apus = [], [], []
    for a in apus:
        info = {"codigo": a.codigo, "turno": a.shift, "nombre": a.nombre,
                "unidad": a.unidad, "grupo": a.grupo,
                "n_componentes": len(comps_por.get((a.codigo, a.shift), []))}
        if alm.apus.get_apu(a.codigo, a.shift):
            ya_existe.append(info)
        else:
            crear.append(info)
            crear_apus.append(a)
    subapus = detectar_subapus_lote(alm, apus, comps_por, solo=crear_apus)
    return {"crear": crear, "ya_existe": ya_existe, "subapus": subapus}


def aplicar_importar_apus(alm: Almacen, contenido: bytes, actor=None) -> dict:
    apus, comps_por = _parse_apus(contenido)
    mapa = mapa_codigos_apu(alm, apus)
    nombres = nombres_apu(alm, apus)
    creados, subapus_marcados, errores = 0, 0, []
    lote = nuevo_lote()
    for a in apus:
        if alm.apus.get_apu(a.codigo, a.shift):
            continue                                   # ya existe: no se pisa
        try:
            comps = comps_por.get((a.codigo, a.shift), [])
            comps, n_sub = marcar_comps_subapu(comps, a.shift, mapa, nombres)
            with alm.transaccion("apus") as conn:
                alm.apus.crear_apu(a, comps, conn=conn)
                registrar_auditoria(
                    alm, conn, actor, "apu.crear", "apu", a.codigo, antes=None,
                    despues={"codigo": a.codigo, "turno": a.shift, "nombre": a.nombre,
                             "unidad": a.unidad, "grupo": a.grupo,
                             "n_componentes": len(comps), "n_subapus": n_sub},
                    contexto={"origen": "import", "lote_id": lote})
            creados += 1
            subapus_marcados += n_sub
        except ValueError as e:
            errores.append({"codigo": a.codigo, "turno": a.shift, "error": str(e)})
    return {"creados": creados, "subapus_marcados": subapus_marcados, "errores": errores}
