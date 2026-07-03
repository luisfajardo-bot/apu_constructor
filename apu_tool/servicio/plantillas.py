"""Generación de plantillas .xlsx para los importadores (APUs, insumos, precios).

Cada plantilla se arma al vuelo con openpyxl usando las MISMAS columnas que espera
el parser correspondiente, de modo que plantilla y parser no puedan desincronizarse
(ver tests/test_plantillas.py: round-trip). NO toca la IA (Invariante #1: catálogo).
"""
from __future__ import annotations

import io

import openpyxl

from apu_tool.datos.seed import APUS_COLS, APUS_SHEET


def _a_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Encabezados de la hoja APUS en las posiciones que dicta APUS_COLS (sin drift).
_APUS_HEADERS = {
    APUS_COLS["actividad"]: "ACTIVIDAD",
    APUS_COLS["cod_idu"]: "COD IDU",
    APUS_COLS["unidad"]: "UN",
    APUS_COLS["insumo_nombre"]: "INSUMO",
    APUS_COLS["insumo_cod"]: "COD",
    APUS_COLS["insumo_und"]: "UND",
    APUS_COLS["rendimiento"]: "RENDIMIENTO",
    APUS_COLS["precio_unitario"]: "PRECIO UNITARIO",
    APUS_COLS["shift"]: "DIURNO/NOCTURNO",
}


def plantilla_apus() -> bytes:
    """Hoja 'APUS': fila de encabezados + 1 APU de ejemplo (encabezado + 2 componentes).

    El COD IDU del ejemplo es numérico ('999001') porque el parser detecta el
    encabezado de un APU con `_looks_like_code`.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = APUS_SHEET
    ancho = max(_APUS_HEADERS) + 1

    fila_h = [""] * ancho
    for i, txt in _APUS_HEADERS.items():
        fila_h[i] = txt
    ws.append(fila_h)

    fila_apu = [""] * ancho
    fila_apu[APUS_COLS["actividad"]] = "EJEMPLO — reemplazar por su actividad"
    fila_apu[APUS_COLS["cod_idu"]] = "999001"
    fila_apu[APUS_COLS["unidad"]] = "M2"
    fila_apu[APUS_COLS["shift"]] = "DIURNO"
    ws.append(fila_apu)

    for nombre, cod, und, rend in [
        ("EJEMPLO cemento gris", "6140", "KG", 0.5),
        ("EJEMPLO arena", "6200", "M3", 0.02),
    ]:
        fila_c = [""] * ancho
        fila_c[APUS_COLS["insumo_nombre"]] = nombre
        fila_c[APUS_COLS["insumo_cod"]] = cod
        fila_c[APUS_COLS["insumo_und"]] = und
        fila_c[APUS_COLS["rendimiento"]] = rend
        ws.append(fila_c)
    return _a_bytes(wb)


def plantilla_licitacion() -> bytes:
    """Plantilla de la lista de licitación (entrada para armar una corrida).

    Columnas: ITEM, DESCRIPCION, UNIDAD, CANTIDAD, PRECIO CONTRACTUAL, TURNO
    (mismas que lee `dominio.licitacion.read_licitacion`). El emparejamiento usa
    la DESCRIPCIÓN; el TURNO debe ser DIURNO o NOCTURNO por ítem.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LICITACION"
    ws.append(["ITEM", "DESCRIPCION", "UNIDAD", "CANTIDAD", "PRECIO CONTRACTUAL", "TURNO"])
    ws.append(["EJEMPLO-1", "EXCAVACIÓN MANUAL EN MATERIAL COMÚN", "M3", 10, 25000, "DIURNO"])
    ws.append(["EJEMPLO-2", "SUMINISTRO E INSTALACIÓN DE SARDINEL A-10", "ML", 5, 40000, "NOCTURNO"])
    return _a_bytes(wb)


def plantilla_insumos() -> bytes:
    """Plantilla del importador unificado. Columnas: codigo, nombre, unidad, grupo,
    precio, fuente. Con nombre crea o actualiza (por identidad código+nombre); sin
    nombre solo actualiza precio por código."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["codigo", "nombre", "unidad", "grupo", "precio", "fuente"])
    ws.append(["EJEMPLO-1", "EJEMPLO — con nombre se crea o actualiza",
               "KG", "MAT", 1000, "COTIZACIÓN"])
    ws.append(["EJEMPLO-2", "", "", "", 2000, "COTIZACIÓN"])  # sin nombre = solo actualizar precio
    return _a_bytes(wb)
