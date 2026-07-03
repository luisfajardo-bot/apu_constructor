"""API de autoría: crear/importar insumos y APUs + listar/detalle de APUs."""
import io
import openpyxl

from apu_tool.datos.almacen import Almacen
from apu_tool.nucleo.models import Apu, Insumo
from apu_tool.servicio.app import create_app
from tests.conftest import cliente

_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _cli(tmp_path):
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([Insumo("100", "CEMENTO GRIS", "KG", "MAT", 1000, "PRECIO IDU")])
    alm.apus.insert_apus([Apu("A1", "MURO EXISTENTE", "M2", "DIURNO", "ESTR")])
    return cliente(create_app(almacen=alm), rol="admin"), alm


def test_crear_insumo_endpoint(tmp_path):
    cli, _ = _cli(tmp_path)
    r = cli.post("/api/insumos/crear", json={"codigo": "300", "nombre": "GRAVA",
                 "unidad": "M3", "grupo": "MAT", "precio": 80000, "fuente": "PRECIO IDU"})
    assert r.status_code == 200, r.text
    assert r.json()["codigo"] == "300"
    # duplicado -> 400
    r2 = cli.post("/api/insumos/crear", json={"codigo": "100", "nombre": "CEMENTO GRIS",
                  "precio": 1})
    assert r2.status_code == 400


def test_crear_apu_endpoint(tmp_path):
    cli, alm = _cli(tmp_path)
    r = cli.post("/api/apus/crear", json={"codigo": "B2", "turno": "DIURNO",
        "nombre": "PISO", "unidad": "M2", "grupo": "ACAB",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    assert r.status_code == 200, r.text
    det = cli.get("/api/apus/B2/DIURNO")
    assert det.status_code == 200
    assert det.json()["composicion"][0]["insumo_codigo"] == "100"
    # duplicado -> 400
    assert cli.post("/api/apus/crear", json={"codigo": "A1", "turno": "DIURNO",
                    "nombre": "MURO"}).status_code == 400


def test_listar_apus_endpoint(tmp_path):
    cli, _ = _cli(tmp_path)
    r = cli.get("/api/apus")
    assert r.status_code == 200
    body = r.json()
    assert body["total"] == 1 and body["items"][0]["codigo"] == "A1"


def _xlsx_insumos() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["codigo", "nombre", "unidad", "grupo", "precio", "fuente"])
    ws.append(["300", "GRAVA COMUN", "M3", "MAT", 80000, "PRECIO IDU"])
    ws.append(["100", "CEMENTO GRIS", "KG", "MAT", 1200, "PRECIO IDU"])  # ya existe
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_import_insumos_endpoint(tmp_path):
    cli, _ = _cli(tmp_path)
    data = _xlsx_insumos()
    pv = cli.post("/api/insumos/importar/preview",
                  files={"archivo": ("insumos.xlsx", data, _XLSX)})
    assert pv.status_code == 200
    assert [c["codigo"] for c in pv.json()["crear"]] == ["300"]
    assert [c["codigo"] for c in pv.json()["actualizar"]] == ["100"]   # existía -> actualizar
    ap = cli.post("/api/insumos/importar",
                  files={"archivo": ("insumos.xlsx", data, _XLSX)})
    assert ap.status_code == 200
    assert ap.json()["creados"] == 1 and ap.json()["actualizados"] == 1


def _xlsx_apus() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "APUS"
    ws.append(["ACT","COD IDU","UN","INSUMO","COD","UND","REND","INV","PRECIO","COSTO","TURNO"])
    ws.append(["MURO NUEVO","7777","M2","","","","","","","","DIURNO"])
    ws.append(["","","","CEMENTO","100","KG",2.5,"",900,"",""])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_import_apus_endpoint(tmp_path):
    cli, alm = _cli(tmp_path)
    data = _xlsx_apus()
    pv = cli.post("/api/apus/importar/preview", files={"archivo": ("apus.xlsx", data, _XLSX)})
    assert pv.status_code == 200
    assert any(c["codigo"] == "7777" for c in pv.json()["crear"])
    ap = cli.post("/api/apus/importar", files={"archivo": ("apus.xlsx", data, _XLSX)})
    assert ap.status_code == 200 and ap.json()["creados"] == 1
    assert alm.apus.get_apu("7777", "DIURNO") is not None


def test_import_apus_archivo_malo_400(tmp_path):
    cli, _ = _cli(tmp_path)
    wb = openpyxl.Workbook(); wb.active.append(["x"]); buf = io.BytesIO(); wb.save(buf)
    r = cli.post("/api/apus/importar/preview",
                 files={"archivo": ("malo.xlsx", buf.getvalue(), _XLSX)})
    assert r.status_code == 400


def test_plantilla_apus_endpoint(tmp_path):
    cli, _ = _cli(tmp_path)
    r = cli.get("/api/apus/importar/plantilla")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX
    assert "attachment" in r.headers["content-disposition"]
    assert len(r.content) > 0
    # lo descargado es re-importable por su propio endpoint (round-trip end-to-end)
    pv = cli.post("/api/apus/importar/preview",
                  files={"archivo": ("plantilla_apus.xlsx", r.content, _XLSX)})
    assert pv.status_code == 200
    assert any(c["codigo"] == "999001" for c in pv.json()["crear"])


def test_plantilla_insumos_endpoint(tmp_path):
    cli, _ = _cli(tmp_path)
    r = cli.get("/api/insumos/importar/plantilla")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX
    assert "attachment" in r.headers["content-disposition"]
    assert len(r.content) > 0


def test_plantilla_licitacion_endpoint(tmp_path):
    cli, _ = _cli(tmp_path)
    r = cli.get("/api/corridas/plantilla")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == _XLSX
    assert "attachment" in r.headers["content-disposition"]
    assert len(r.content) > 0


def test_plantillas_requieren_editor(tmp_path):
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    cli = cliente(create_app(almacen=alm), rol="consulta")
    assert cli.get("/api/apus/importar/plantilla").status_code == 403
    assert cli.get("/api/insumos/importar/plantilla").status_code == 403
    # la de licitación es rol consulta (cualquiera que arma corridas puede bajarla)
    assert cli.get("/api/corridas/plantilla").status_code == 200


def test_detalle_apu_incluye_n_corridas(tmp_path):
    cli, alm = _cli(tmp_path)   # A1 existe por el seed de _cli; sin corridas -> 0
    r = cli.get("/api/apus/A1/DIURNO")
    assert r.status_code == 200, r.text
    assert r.json()["n_corridas"] == 0


def test_editar_apu_endpoint(tmp_path):
    cli, alm = _cli(tmp_path)   # rol admin
    cli.post("/api/apus/crear", json={"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "unidad": "M2", "grupo": "ACAB",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    r = cli.put("/api/apus/B2/DIURNO", json={"nombre": "PISO PULIDO", "unidad": "M2",
        "grupo": "ACAB", "componentes": [{"insumo_codigo": "100", "rendimiento": 3.0}]})
    assert r.status_code == 200, r.text
    assert r.json()["nombre"] == "PISO PULIDO"
    assert cli.put("/api/apus/NOPE/DIURNO", json={"nombre": "X",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 1.0}]}).status_code == 404


def test_borrar_apu_endpoint(tmp_path):
    cli, alm = _cli(tmp_path)   # rol admin
    cli.post("/api/apus/crear", json={"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    r = cli.delete("/api/apus/B2/DIURNO")
    assert r.status_code == 200 and r.json()["borrado"] is True
    assert cli.delete("/api/apus/NOPE/DIURNO").status_code == 404


def test_editar_borrar_gating_por_rol(tmp_path):
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([Insumo("100", "CEMENTO GRIS", "KG", "MAT", 1000, "PRECIO IDU")])
    alm.apus.insert_apus([Apu("A1", "MURO", "M2", "DIURNO", "ESTR")])
    editor = cliente(create_app(almacen=alm), rol="editor")
    consulta = cliente(create_app(almacen=alm), rol="consulta")
    body = {"nombre": "MURO 2", "unidad": "M2", "grupo": "ESTR",
            "componentes": [{"insumo_codigo": "100", "rendimiento": 1.0}]}
    assert editor.put("/api/apus/A1/DIURNO", json=body).status_code == 200   # editor edita
    assert editor.delete("/api/apus/A1/DIURNO").status_code == 403           # editor NO borra
    assert consulta.put("/api/apus/A1/DIURNO", json=body).status_code == 403 # consulta NO edita
