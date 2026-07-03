"""Plantillas de importación: round-trip contra su propio parser (candado anti-drift)."""
import io

import openpyxl

from apu_tool.datos.almacen import Almacen
from apu_tool.servicio import plantillas


def _alm(tmp_path):
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    return alm


def test_plantillas_abren_como_workbook_valido():
    for gen in (plantillas.plantilla_apus, plantillas.plantilla_insumos_crear,
                plantillas.plantilla_precios):
        data = gen()
        assert data, f"{gen.__name__} devolvió vacío"
        wb = openpyxl.load_workbook(io.BytesIO(data))
        wb.close()
    wb = openpyxl.load_workbook(io.BytesIO(plantillas.plantilla_apus()))
    assert "APUS" in wb.sheetnames  # el parser exige la hoja 'APUS'
    wb.close()


def test_plantilla_apus_round_trip(tmp_path):
    from apu_tool.servicio import autoria
    alm = _alm(tmp_path)
    pv = autoria.preview_importar_apus(alm, plantillas.plantilla_apus())
    apus = pv["crear"] + pv["ya_existe"]
    assert len(apus) == 1
    assert apus[0]["codigo"] == "999001"
    assert apus[0]["n_componentes"] == 2


def test_plantilla_insumos_round_trip(tmp_path):
    from apu_tool.servicio import autoria
    alm = _alm(tmp_path)
    pv = autoria.preview_importar_insumos(alm, plantillas.plantilla_insumos_crear(),
                                          "plantilla_insumos.xlsx")
    codigos = [f["codigo"] for f in pv["crear"]]
    assert "EJEMPLO-1" in codigos


def test_plantilla_precios_round_trip():
    from apu_tool.servicio.insumos import _parse_tabla
    filas = _parse_tabla(plantillas.plantilla_precios(), "plantilla_precios.xlsx")
    assert len(filas) == 1
    assert filas[0]["codigo"] == "EJEMPLO-1"
    assert filas[0]["precio"] == 1000.0
