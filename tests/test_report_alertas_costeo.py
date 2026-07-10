import openpyxl
from apu_tool.dominio.report import write_report, _ALERT_FILL, _BAD_FILL
from apu_tool.nucleo.models import AssembledApu, CostedComponent, LicitacionItem, MatchStatus


def _item(item="1"):
    return LicitacionItem(item=item, descripcion="Losa", unidad="m3", cantidad=2.0,
                          precio_contractual=100.0, shift="DIURNO")


def _apu(comps, costo, status=MatchStatus.AUTO):
    return AssembledApu(item=_item(), apu_codigo="A", apu_nombre="Losa", unidad="m3",
                        shift="DIURNO", componentes=comps, costo_unitario=costo,
                        status=status, confianza=1.0)


def _comp(costo, precio, calidad="exacto"):
    return CostedComponent(insumo_codigo="7", insumo_nombre="Cemento", unidad="kg",
                           rendimiento=1.0, precio_unitario=precio, fuente_precio="X",
                           costo=costo, calidad_cruce=calidad)


def test_resumen_resalta_fila_con_costeo_en_cero(tmp_path):
    out = write_report([_apu([_comp(0.0, 0.0)], 0.0)], tmp_path / "c.xlsx")
    ws = openpyxl.load_workbook(out)["RESUMEN"]
    # fila 2 = primer ítem; celda con fill de alerta de costeo
    assert ws.cell(row=2, column=1).fill.fgColor.rgb.endswith("F8CBAD")


def test_alertas_lista_motivo_de_costeo(tmp_path):
    out = write_report([_apu([_comp(0.0, 0.0)], 0.0)], tmp_path / "c.xlsx")
    ws = openpyxl.load_workbook(out)["ALERTAS"]
    textos = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
    assert any(t and "en $0" in t for t in textos)


def test_resumen_prioridad_alerta_costeo_sobre_margen_negativo(tmp_path):
    # Ítem 1: componente en $0 (alerta de costeo) Y margen negativo -> gana _ALERT_FILL.
    apu_con_alerta = _apu([_comp(0.0, 0.0)], costo=150.0)
    # Ítem 2: solo margen negativo, sin alerta de costeo -> _BAD_FILL.
    apu_solo_margen = _apu([_comp(300.0, 300.0)], costo=150.0)
    out = write_report([apu_con_alerta, apu_solo_margen], tmp_path / "c.xlsx")
    ws = openpyxl.load_workbook(out)["RESUMEN"]
    assert ws.cell(row=2, column=1).fill.fgColor.rgb.endswith("F8CBAD")
    assert ws.cell(row=3, column=1).fill.fgColor.rgb.endswith("FCE4E4")
