"""Pruebas del reporte categorizado: agrupación y subtotales por capítulo."""
import openpyxl
import pytest

from apu_tool.nucleo.models import AssembledApu, LicitacionItem, MatchStatus, CostedComponent
from apu_tool.dominio.report_categorizado import agrupar_por_capitulo, write_report_categorizado


def _apu(cat, item, contractual, costo_unit, cant=1.0):
    it = LicitacionItem(item=item, descripcion=f"ACT {item}", unidad="M3",
                        cantidad=cant, precio_contractual=contractual, shift="DIURNO",
                        categoria=cat, codigo_sugerido=item)
    comp = CostedComponent("100", "CEMENTO", "KG", 3.0, costo_unit / 3.0,
                           "PRECIO IDU", costo_unit)
    return AssembledApu(item=it, apu_codigo=item, apu_nombre=f"ACT {item}",
                        unidad="M3", shift="DIURNO", componentes=[comp],
                        costo_unitario=costo_unit, status=MatchStatus.AUTO, confianza=1.0)


def test_agrupar_por_capitulo_preserva_orden():
    apus = [_apu("A", "1", 100, 80), _apu("B", "2", 200, 150), _apu("A", "3", 50, 40)]
    grupos = agrupar_por_capitulo(apus)
    assert list(grupos.keys()) == ["A", "B"]
    assert len(grupos["A"]) == 2 and len(grupos["B"]) == 1


def test_reporte_subtotales_y_gran_total(tmp_path):
    apus = [_apu("A", "1", 100, 80, cant=2.0),   # contractual 200, costo 160
            _apu("B", "2", 200, 150, cant=1.0)]   # contractual 200, costo 150
    out = write_report_categorizado(apus, tmp_path / "cat.xlsx")
    wb = openpyxl.load_workbook(out, data_only=True)
    assert "RESUMEN POR CAPÍTULO" in wb.sheetnames
    assert "DETALLE" in wb.sheetnames
    assert "APUS" in wb.sheetnames

    # En RESUMEN, debe existir una fila de gran total con contractual 400 y costo 310.
    ws = wb["RESUMEN POR CAPÍTULO"]
    valores = [tuple(r) for r in ws.iter_rows(values_only=True)]
    plano = [c for fila in valores for c in fila]
    assert 400 in plano       # gran total contractual
    assert 310 in plano       # gran total costo


def test_columna_cruce_aparece_en_el_detalle(tmp_path):
    it = LicitacionItem(item="1", descripcion="ACT 1", unidad="M3", cantidad=1.0,
                        precio_contractual=100, shift="DIURNO", categoria="A",
                        codigo_sugerido="1")
    comp = CostedComponent("4513", "BASE GRANULAR", "M3", 1.0, 190300, "PRECIO IDU",
                           190300, calidad_cruce="aproximado")
    apu = AssembledApu(item=it, apu_codigo="1", apu_nombre="ACT 1", unidad="M3",
                       shift="DIURNO", componentes=[comp], costo_unitario=190300,
                       status=MatchStatus.AUTO, confianza=1.0)
    out = write_report_categorizado([apu], tmp_path / "cat.xlsx")
    wb = openpyxl.load_workbook(out, data_only=True)
    celdas = [c for ws in wb.worksheets for fila in ws.iter_rows(values_only=True)
              for c in fila]
    assert "Cruce" in celdas        # encabezado de la columna
    assert "aproximado" in celdas   # el aviso del cruce, en la fila del insumo
