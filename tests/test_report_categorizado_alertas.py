import openpyxl
from apu_tool.dominio.report_categorizado import write_report_categorizado
from apu_tool.nucleo.models import AssembledApu, CostedComponent, LicitacionItem, MatchStatus


def _apu():
    item = LicitacionItem(item="1", descripcion="Losa", unidad="m3", cantidad=1.0,
                          precio_contractual=100.0, shift="DIURNO", categoria="CAP-1")
    comp = CostedComponent(insumo_codigo="7", insumo_nombre="Cemento", unidad="kg",
                           rendimiento=1.0, precio_unitario=0.0, fuente_precio="X",
                           costo=0.0, calidad_cruce="exacto")
    return AssembledApu(item=item, apu_codigo="A", apu_nombre="Losa", unidad="m3",
                        shift="DIURNO", componentes=[comp], costo_unitario=0.0,
                        status=MatchStatus.AUTO, confianza=1.0)


def test_detalle_resalta_costeo_cero(tmp_path):
    out = write_report_categorizado([_apu()], tmp_path / "c.xlsx")
    ws = openpyxl.load_workbook(out)["DETALLE"]
    tiene_alerta = any(ws.cell(row=r, column=1).fill.fgColor.rgb.endswith("F8CBAD")
                       for r in range(1, ws.max_row + 1))
    assert tiene_alerta


def test_alertas_incluye_costeo(tmp_path):
    out = write_report_categorizado([_apu()], tmp_path / "c.xlsx")
    ws = openpyxl.load_workbook(out)["ALERTAS"]
    textos = [ws.cell(row=r, column=7).value for r in range(2, ws.max_row + 1)]
    assert any(t and "en $0" in t for t in textos)
