"""El cuadro documenta con qué distancias y ajustes se costeó."""
import openpyxl

from apu_tool.dominio.report import write_report
from apu_tool.nucleo.models import (
    AjusteProyecto, AssembledApu, CostedComponent, LicitacionItem, MatchStatus,
    ParametrosProyecto)


def _ens():
    item = LicitacionItem(item="1", descripcion="RELLENO", unidad="M3", cantidad=2,
                          precio_contractual=100000.0, shift="DIURNO")
    comp = CostedComponent(insumo_codigo="7462", insumo_nombre="TRANSPORTE DE PETREOS",
                           unidad="M3-KM", rendimiento=33.6, precio_unitario=1000.0,
                           fuente_precio="COSTO INTERNO", costo=33600)
    return AssembledApu(item=item, apu_codigo="4390", apu_nombre="RELLENO",
                        unidad="M3", shift="DIURNO", componentes=[comp],
                        costo_unitario=33600, status=MatchStatus.CONFIRMED,
                        confianza=1.0)


def test_sin_desviaciones_no_hay_hoja(tmp_path):
    out = write_report([_ens()], tmp_path / "cuadro.xlsx")
    wb = openpyxl.load_workbook(out)
    assert "DESVIACIONES DEL PROYECTO" not in wb.sheetnames


def test_con_desviaciones_la_hoja_las_lista(tmp_path):
    out = write_report(
        [_ens()], tmp_path / "cuadro.xlsx",
        parametros=ParametrosProyecto(km_botadero=34, km_mezclas=28, km_granulares=32,
                                      peaje_aplica=True, peaje_valor=12400),
        ajustes=[AjusteProyecto(apu_codigo="4390", shift="DIURNO", accion="agregar",
                                insumo_codigo="9001",
                                insumo_nombre="GEOTEXTIL NT 2000", unidad="M2",
                                rendimiento=1.1, nota="especificación del cliente")])
    wb = openpyxl.load_workbook(out)
    ws = wb["DESVIACIONES DEL PROYECTO"]
    texto = "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                      if c.value is not None)
    assert "34" in texto and "Botadero" in texto
    assert "12400" in texto or "12.400" in texto
    assert "GEOTEXTIL NT 2000" in texto and "especificación del cliente" in texto
