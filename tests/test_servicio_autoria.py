"""Servicio de autoría de base: crear insumos/APUs (individual + Excel)."""
import io
import openpyxl
import pytest

from apu_tool.datos.almacen import Almacen
from apu_tool.nucleo.models import Apu, ApuComponent, Insumo
from apu_tool.servicio import autoria


def _alm(tmp_path):
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([
        Insumo("100", "CEMENTO GRIS", "KG", "MAT", 1000, "PRECIO IDU"),
        Insumo("200", "ARENA", "M3", "MAT", 50000, "PRECIO IDU")])
    alm.apus.insert_apus([Apu("A1", "MURO EXISTENTE", "M2", "DIURNO", "ESTR")])
    return alm


# ---------------------------------------------------------------- individual
def test_crear_insumo_ok(tmp_path):
    alm = _alm(tmp_path)
    out = autoria.crear_insumo(alm, {"codigo": "300", "nombre": "GRAVA", "unidad": "M3",
                                     "grupo": "MAT", "precio": 80000, "fuente": "PRECIO IDU"})
    assert out["codigo"] == "300" and out["precio"] == 80000
    assert any(i.codigo == "300" for i in alm.precios.get_candidatos("300"))


def test_crear_insumo_duplicado_y_validacion(tmp_path):
    alm = _alm(tmp_path)
    with pytest.raises(ValueError):
        autoria.crear_insumo(alm, {"codigo": "100", "nombre": "CEMENTO GRIS", "precio": 1})
    with pytest.raises(ValueError):
        autoria.crear_insumo(alm, {"codigo": "", "nombre": "X", "precio": 1})
    with pytest.raises(ValueError):
        autoria.crear_insumo(alm, {"codigo": "9", "nombre": "X", "precio": -5})


def test_crear_apu_con_composicion(tmp_path):
    alm = _alm(tmp_path)
    out = autoria.crear_apu(alm, {"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "unidad": "M2", "grupo": "ACAB",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0},
                        {"insumo_codigo": "200", "rendimiento": 0.5}]})
    assert out["codigo"] == "B2" and out["n_componentes"] == 2
    comps = alm.apus.get_components("B2", "DIURNO")
    # nombre/unidad se resolvieron desde la base
    assert comps[0].insumo_nombre == "CEMENTO GRIS" and comps[0].unidad == "KG"


def test_crear_apu_validaciones(tmp_path):
    alm = _alm(tmp_path)
    with pytest.raises(ValueError):  # turno inválido
        autoria.crear_apu(alm, {"codigo": "Z", "turno": "TARDE", "nombre": "X"})
    with pytest.raises(ValueError):  # rendimiento <= 0
        autoria.crear_apu(alm, {"codigo": "Z", "turno": "DIURNO", "nombre": "X",
            "componentes": [{"insumo_codigo": "100", "rendimiento": 0}]})
    with pytest.raises(ValueError):  # duplicado (A1, DIURNO)
        autoria.crear_apu(alm, {"codigo": "A1", "turno": "DIURNO", "nombre": "MURO"})


# ---------------------------------------------------------------- import insumos
def _xlsx_upsert() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["codigo", "nombre", "unidad", "grupo", "precio", "fuente"])
    ws.append(["300", "GRAVA COMUN", "M3", "MAT", 80000, "PRECIO IDU"])   # con nombre, no existe -> crear
    ws.append(["100", "CEMENTO GRIS", "KG", "MAT", 1200, "PRECIO IDU"])   # con nombre, existe -> actualizar
    ws.append(["", "SIN CODIGO", "UN", "", 10, ""])                       # sin codigo -> invalida
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def _xlsx_solo_precio(filas) -> bytes:
    """Archivo estilo lista de precios: codigo, precio (sin nombre)."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.append(["codigo", "precio", "fuente"])
    for f in filas:
        ws.append(f)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_upsert_preview_con_nombre(tmp_path):
    alm = _alm(tmp_path)
    prev = autoria.preview_importar_insumos(alm, _xlsx_upsert(), "insumos.xlsx")
    assert [c["codigo"] for c in prev["crear"]] == ["300"]
    assert [c["codigo"] for c in prev["actualizar"]] == ["100"]
    assert prev["actualizar"][0]["precio_actual"] == 1000 and prev["actualizar"][0]["precio_nuevo"] == 1200
    assert len(prev["invalida"]) == 1


def test_upsert_aplicar_crea_y_actualiza(tmp_path):
    alm = _alm(tmp_path)
    res = autoria.aplicar_importar_insumos(alm, _xlsx_upsert(), "insumos.xlsx")
    assert res["creados"] == 1 and res["actualizados"] == 1
    assert any(i.codigo == "300" for i in alm.precios.get_candidatos("300"))
    assert alm.precios.get_candidatos("100")[0].precio == 1200   # precio actualizado


def test_upsert_sin_nombre_codigo_unico_actualiza(tmp_path):
    alm = _alm(tmp_path)
    prev = autoria.preview_importar_insumos(alm, _xlsx_solo_precio([["100", 1500, "COMPRAS"]]),
                                            "precios.xlsx")
    assert len(prev["actualizar"]) == 1 and prev["actualizar"][0]["precio_nuevo"] == 1500
    assert prev["crear"] == [] and prev["no_encontrada"] == []


def test_upsert_sin_nombre_codigo_repetido_ambiguo(tmp_path):
    alm = _alm(tmp_path)
    alm.precios.insert_insumos([
        Insumo("100", "CEMENTO BLANCO", "KG", "MAT", 2000, "PRECIO IDU")])
    prev = autoria.preview_importar_insumos(alm, _xlsx_solo_precio([["100", 1500, "X"]]),
                                            "precios.xlsx")
    assert len(prev["ambigua"]) == 1 and prev["ambigua"][0]["codigo"] == "100"
    assert len(prev["ambigua"][0]["candidatos"]) == 2


def test_upsert_sin_nombre_codigo_inexistente_no_encontrada(tmp_path):
    alm = _alm(tmp_path)
    prev = autoria.preview_importar_insumos(alm, _xlsx_solo_precio([["999", 1500, "X"]]),
                                            "precios.xlsx")
    assert [n["codigo"] for n in prev["no_encontrada"]] == ["999"]


def test_upsert_precio_vacio_en_actualizacion_no_cambia(tmp_path):
    alm = _alm(tmp_path)
    prev = autoria.preview_importar_insumos(alm, _xlsx_solo_precio([["100", "", "NUEVA FUENTE"]]),
                                            "precios.xlsx")
    c = prev["actualizar"][0]
    assert c["precio_nuevo"] == 1000            # precio actual, no 0
    assert c["fuente_nueva"] == "NUEVA FUENTE"  # la fuente sí se cambia


# ---------------------------------------------------------------- import APUs
def _xlsx_apus() -> bytes:
    wb = openpyxl.Workbook(); ws = wb.active
    # formato hoja APUS: actividad(0) cod_idu(1) unidad(2) insumo(3) cod(4) und(5)
    #                    rendimiento(6) inv(7) precio(8) costo(9) turno(10)
    ws.title = "APUS"
    ws.append(["ACTIVIDAD","COD IDU","UN","INSUMO","COD","UND","RENDIMIENTO","INV","PRECIO","COSTO","TURNO"])
    ws.append(["MURO NUEVO ESPECIAL","7777","M2","","","","","","","","DIURNO"])  # cabecera APU
    ws.append(["","","","CEMENTO","100","KG",2.5,"",900,"",""])                   # componente
    ws.append(["","","","ARENA","200","M3",0.5,"",50,"",""])                      # componente
    ws.append(["MURO EXISTENTE","A1","M2","","","","","","","","DIURNO"])         # ya existe
    ws.append(["","","","CEMENTO","100","KG",1.0,"",900,"",""])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_import_apus_preview_y_aplicar(tmp_path):
    alm = _alm(tmp_path)
    data = _xlsx_apus()
    prev = autoria.preview_importar_apus(alm, data)
    codigos_crear = {c["codigo"] for c in prev["crear"]}
    assert "7777" in codigos_crear
    assert any(c["codigo"] == "A1" for c in prev["ya_existe"])
    nuevo = next(c for c in prev["crear"] if c["codigo"] == "7777")
    assert nuevo["n_componentes"] == 2 and nuevo["turno"] == "DIURNO"
    res = autoria.aplicar_importar_apus(alm, data)
    assert res["creados"] == 1
    assert alm.apus.get_apu("7777", "DIURNO").nombre == "MURO NUEVO ESPECIAL"
    assert len(alm.apus.get_components("7777", "DIURNO")) == 2


def _xlsx_apus_nocturno() -> bytes:
    """Plantilla con nocturnos de código PELADO (la N va solo en la columna turno),
    como en la lista de licitación real que rompió en prod."""
    wb = openpyxl.Workbook(); ws = wb.active
    ws.title = "APUS"
    ws.append(["ACTIVIDAD","COD IDU","UN","INSUMO","COD","UND","RENDIMIENTO","INV","PRECIO","COSTO","TURNO"])
    ws.append(["EXCAVACION NOCTURNA","8888","M3","","","","","","","","NOCTURNO"])   # pelado -> debe quedar "8888 N"
    ws.append(["","","","CEMENTO","100","KG",1.5,"",900,"",""])
    ws.append(["DEMOLICION DIURNA","9999","M2","","","","","","","","DIURNO"])       # diurno -> intacto
    ws.append(["","","","ARENA","200","M3",1.0,"",50,"",""])
    ws.append(["YA CON N","7000 N","M3","","","","","","","","NOCTURNO"])            # ya trae N -> idempotente
    ws.append(["","","","CEMENTO","100","KG",1.0,"",900,"",""])
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


def test_import_apus_nocturno_agrega_sufijo_n(tmp_path):
    alm = _alm(tmp_path)
    data = _xlsx_apus_nocturno()
    prev = autoria.preview_importar_apus(alm, data)
    codigos = {(c["codigo"], c["turno"]) for c in prev["crear"]}
    assert ("8888 N", "NOCTURNO") in codigos       # nocturno pelado -> con sufijo
    assert ("8888", "NOCTURNO") not in codigos     # no se crea el pelado
    assert ("9999", "DIURNO") in codigos           # diurno intacto
    assert ("7000 N", "NOCTURNO") in codigos       # ya tenía N -> sin doble sufijo
    assert ("7000 N N", "NOCTURNO") not in codigos

    res = autoria.aplicar_importar_apus(alm, data)
    assert res["creados"] == 3
    assert alm.apus.get_apu("8888 N", "NOCTURNO") is not None
    assert alm.apus.get_apu("8888", "NOCTURNO") is None
    assert len(alm.apus.get_components("8888 N", "NOCTURNO")) == 1   # el componente cuelga del código con N
    assert alm.apus.get_apu("7000 N", "NOCTURNO") is not None
    assert alm.apus.get_apu("9999", "DIURNO") is not None


def test_import_apus_sin_hoja_apus(tmp_path):
    alm = _alm(tmp_path)
    wb = openpyxl.Workbook(); wb.active.append(["x"]); buf = io.BytesIO(); wb.save(buf)
    with pytest.raises(ValueError):
        autoria.preview_importar_apus(alm, buf.getvalue())


def test_editar_apu_reemplaza_y_devuelve_resumen(tmp_path):
    alm = _alm(tmp_path)
    autoria.crear_apu(alm, {"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "unidad": "M2", "grupo": "ACAB",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    out = autoria.editar_apu(alm, "B2", "DIURNO", {"nombre": "PISO PULIDO",
        "unidad": "M2", "grupo": "ACAB",
        "componentes": [{"insumo_codigo": "200", "rendimiento": 0.5}]})
    assert out["nombre"] == "PISO PULIDO" and out["n_componentes"] == 1
    comps = alm.apus.get_components("B2", "DIURNO")
    assert [c.insumo_codigo for c in comps] == ["200"]


def test_editar_apu_inexistente_devuelve_none(tmp_path):
    alm = _alm(tmp_path)
    assert autoria.editar_apu(alm, "NOPE", "DIURNO", {"nombre": "X",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 1.0}]}) is None


def test_editar_apu_rendimiento_invalido_lanza(tmp_path):
    alm = _alm(tmp_path)
    autoria.crear_apu(alm, {"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    with pytest.raises(ValueError):
        autoria.editar_apu(alm, "B2", "DIURNO", {"nombre": "PISO",
            "componentes": [{"insumo_codigo": "100", "rendimiento": 0}]})


def test_borrar_apu_ok_devuelve_resultado(tmp_path):
    alm = _alm(tmp_path)
    autoria.crear_apu(alm, {"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    out = autoria.borrar_apu(alm, "B2", "DIURNO")
    assert out == {"borrado": True, "n_corridas": 0}
    assert alm.apus.get_apu("B2", "DIURNO") is None


def test_borrar_apu_inexistente_devuelve_none(tmp_path):
    alm = _alm(tmp_path)
    assert autoria.borrar_apu(alm, "NOPE", "DIURNO") is None


# --------------------------------------------------- FIX 1: preservar tipo/ref_shift
def test_editar_apu_preserva_marca_subapu_si_no_viene_tipo(tmp_path):
    alm = _alm(tmp_path)
    autoria.crear_apu(alm, {"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "unidad": "M2", "grupo": "ACAB",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    alm.apus.set_componente_subapu("B2", "DIURNO", 0, "DIURNO")
    comps = alm.apus.get_components("B2", "DIURNO")
    assert comps[0].tipo == "apu" and comps[0].ref_shift == "DIURNO"

    autoria.editar_apu(alm, "B2", "DIURNO", {"nombre": "PISO PULIDO",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})  # sin 'tipo'
    comps = alm.apus.get_components("B2", "DIURNO")
    assert comps[0].tipo == "apu" and comps[0].ref_shift == "DIURNO"    # preservado


def test_editar_apu_tipo_explicito_gana_sobre_marca_previa(tmp_path):
    alm = _alm(tmp_path)
    autoria.crear_apu(alm, {"codigo": "B2", "turno": "DIURNO", "nombre": "PISO",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0}]})
    alm.apus.set_componente_subapu("B2", "DIURNO", 0, "DIURNO")

    autoria.editar_apu(alm, "B2", "DIURNO", {"nombre": "PISO PULIDO",
        "componentes": [{"insumo_codigo": "100", "rendimiento": 2.0, "tipo": "insumo"}]})
    comps = alm.apus.get_components("B2", "DIURNO")
    assert comps[0].tipo == "insumo" and comps[0].ref_shift == ""      # explícito gana
