"""Las corridas costean con los parámetros de SU proyecto (carpeta raíz)."""
from apu_tool.datos.almacen import Almacen
from apu_tool.dominio import transporte
from apu_tool.nucleo.models import (
    AjusteProyecto, Apu, ApuComponent, ClaseTransporte, Insumo, LicitacionItem,
    ParametrosProyecto)
from apu_tool.servicio import corridas as svc


def _alm(tmp_path):
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([
        Insumo(codigo="7462", nombre="TRANSPORTE DE PETREOS", unidad="M3-KM",
               grupo="TRANSPORTES", precio=1000.0, fuente_precio="COSTO INTERNO")])
    alm.apus.insert_apus([Apu(codigo="4390", nombre="RELLENO", unidad="M3",
                              shift="DIURNO", grupo="VIAS")])
    alm.apus.insert_components([
        ApuComponent(apu_codigo="4390", shift="DIURNO", insumo_codigo="7462",
                     insumo_nombre="TRANSPORTE DE PETREOS", unidad="M3-KM",
                     rendimiento=26.25, precio_unitario_hist=1000.0)])
    alm.apus.set_clasificacion_transporte([ClaseTransporte(
        apu_codigo="4390", shift="DIURNO", insumo_codigo="7462",
        insumo_nombre="TRANSPORTE DE PETREOS", categoria="granulares",
        volumen=1.05, km_base=25.0)])
    return alm


def _corrida(alm, carpeta_id):
    items = [LicitacionItem(item="1", descripcion="RELLENO", unidad="M3", cantidad=10,
                            precio_contractual=100000.0, shift="DIURNO")]
    cid = svc.construir_corrida(alm, "lic.xlsx", items, "DIURNO", False,
                                carpeta_id=carpeta_id)
    svc.confirmar_item(alm, cid, 0, "4390", "DIURNO")
    return cid


def test_cada_proyecto_costea_con_su_distancia(tmp_path):
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    calle13 = alm.carpetas.crear("Calle 13")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=calle13, km_granulares=25))
    c_metro, c_c13 = _corrida(alm, metro), _corrida(alm, calle13)
    v_metro = svc.vista_corrida(alm, c_metro)["items"][0]
    v_c13 = svc.vista_corrida(alm, c_c13)["items"][0]
    assert v_metro["costo_unitario"] == 33600      # 1.05 * 32 * 1000
    assert v_c13["costo_unitario"] == 26250       # 1.05 * 25 * 1000
    # y la biblioteca no cambió
    comps = alm.apus.get_components("4390", "DIURNO")
    assert comps[0].rendimiento == 26.25


def test_subcarpeta_hereda_del_proyecto(tmp_path):
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    lote = alm.carpetas.crear("Lote 2", parent_id=metro)
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    cid = _corrida(alm, lote)
    assert svc.vista_corrida(alm, cid)["items"][0]["costo_unitario"] == 33600


def test_congelada_conserva_su_foto(tmp_path):
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    cid = _corrida(alm, metro)
    svc.congelar(alm, cid)
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=50))
    assert svc.vista_corrida(alm, cid)["items"][0]["costo_unitario"] == 33600


def test_sin_parametros_costea_como_siempre(tmp_path):
    alm = _alm(tmp_path)
    cid = _corrida(alm, alm.carpetas.crear("Sin distancias"))
    assert svc.vista_corrida(alm, cid)["items"][0]["costo_unitario"] == 26250


def test_cargar_contexto_sube_a_la_raiz(tmp_path):
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    lote = alm.carpetas.crear("Lote 2", parent_id=metro)
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_botadero=34))
    ctx = transporte.cargar_contexto(alm, lote)
    assert ctx.params.km_botadero == 34
    assert transporte.cargar_contexto(alm, None).vacio is True


def test_detalle_item_y_cuadro_usan_el_contexto(tmp_path):
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    cid = _corrida(alm, metro)
    det = svc.detalle_item(alm, cid, 0)
    assert det["composicion"][0]["rendimiento"] == 33.6
    assert svc.generar_cuadro(alm, cid) is not None


def test_el_armado_en_vivo_costea_con_las_distancias_del_proyecto(tmp_path):
    """Los eventos del armado tienen que traer el mismo costo que la vista: si no,
    el numero salta cuando termina de armar."""
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    items = [LicitacionItem(item="1", descripcion="RELLENO", unidad="M3", cantidad=10,
                            precio_contractual=100000.0, shift="DIURNO")]
    filas = []
    for evento, payload in svc.construir_corrida_stream(
            alm, "lic.xlsx", items, "DIURNO", False, carpeta_id=metro):
        if evento == "progress":
            filas.append(payload["fila"])
    costeadas = [f for f in filas if f.get("costo_unitario")]
    assert costeadas, filas
    assert all(f["costo_unitario"] == 33600 for f in costeadas)


def test_listar_corridas_resuelve_el_contexto_una_vez_por_proyecto(tmp_path, monkeypatch):
    """El contexto se resuelve por PROYECTO, no por corrida: contra Postgres cada
    resolucion son varios round-trips."""
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    _corrida(alm, metro)
    _corrida(alm, metro)
    _corrida(alm, metro)
    llamadas = []
    real = transporte.cargar_contexto
    def espia(almacen, carpeta_id):
        llamadas.append(carpeta_id)
        return real(almacen, carpeta_id)
    monkeypatch.setattr(svc.transporte, "cargar_contexto", espia)
    svc.listar_corridas(alm)
    assert llamadas.count(metro) == 1, llamadas


def test_la_vista_de_la_corrida_dice_con_que_distancias_costeo(tmp_path):
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32,
                                                  peaje_aplica=True, peaje_valor=12400))
    cid = _corrida(alm, metro)
    meta = svc.vista_corrida(alm, cid)
    assert meta["carpeta_id"] == metro
    assert meta["transporte"]["km_granulares"] == 32
    assert meta["transporte"]["peaje_valor"] == 12400
    # una corrida sin proyecto no trae distancias
    otra = _corrida(alm, alm.carpetas.crear("Sin distancias"))
    assert svc.vista_corrida(alm, otra)["transporte"] is None


def test_la_corrida_alerta_el_componente_sin_clasificar(tmp_path):
    alm = _alm(tmp_path)
    # Se borra la clasificación para simular un APU nuevo sin clasificar.
    with alm.apus.connect() as conn:
        conn.execute("DELETE FROM componente_transporte")
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    cid = _corrida(alm, metro)
    alertas = svc.vista_corrida(alm, cid)["items"][0]["alertas_costeo"]
    assert any("distancia del proyecto no aplicada" in a for a in alertas)


def _alm_con_subapu(tmp_path):
    """Como `_alm`, pero el 4390 usa el sub-APU 3017 (botadero) además de su propio
    transporte: la distancia al botadero vive DENTRO del sub-APU."""
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([
        Insumo(codigo="7462", nombre="TRANSPORTE DE PETREOS", unidad="M3-KM",
               grupo="TRANSPORTES", precio=1000.0, fuente_precio="COSTO INTERNO"),
        Insumo(codigo="7231", nombre="DERECHOS DE BOTADERO", unidad="M3",
               grupo="TRANSPORTES", precio=5000.0, fuente_precio="COSTO INTERNO")])
    alm.apus.insert_apus([
        Apu(codigo="4390", nombre="RELLENO", unidad="M3", shift="DIURNO", grupo="VIAS"),
        Apu(codigo="3017", nombre="TRANSPORTE Y DISPOSICION FINAL DE ESCOMBROS",
            unidad="M3", shift="DIURNO", grupo="TRANSPORTES")])
    alm.apus.insert_components([
        ApuComponent(apu_codigo="4390", shift="DIURNO", insumo_codigo="3017",
                     insumo_nombre="TRANSPORTE Y DISPOSICION FINAL DE ESCOMBROS",
                     unidad="M3", rendimiento=1.0, precio_unitario_hist=20000.0,
                     tipo="apu", ref_shift="DIURNO"),
        ApuComponent(apu_codigo="3017", shift="DIURNO", insumo_codigo="7231",
                     insumo_nombre="DERECHOS DE BOTADERO", unidad="M3",
                     rendimiento=1.3, precio_unitario_hist=5000.0),
        # OJO: este componente queda SIN clasificar a propósito.
        ApuComponent(apu_codigo="3017", shift="DIURNO", insumo_codigo="7462",
                     insumo_nombre="TRANSPORTE DE PETREOS", unidad="M3-KM",
                     rendimiento=20.0, precio_unitario_hist=1000.0)])
    return alm


def test_alerta_del_pendiente_de_un_subapu_menciona_el_subapu(tmp_path):
    """Corrida en un proyecto con km_botadero definido y el acarreo del sub-APU sin
    clasificar: el ítem tiene que alertar mencionando el sub-APU donde vive."""
    alm = _alm_con_subapu(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_botadero=34))
    items = [LicitacionItem(item="1", descripcion="RELLENO", unidad="M3", cantidad=1,
                            precio_contractual=100000.0, shift="DIURNO")]
    cid = svc.construir_corrida(alm, "lic.xlsx", items, "DIURNO", False, carpeta_id=metro)
    svc.confirmar_item(alm, cid, 0, "4390", "DIURNO")
    alertas = svc.vista_corrida(alm, cid)["items"][0]["alertas_costeo"]
    assert any("distancia del proyecto no aplicada" in a and "sub-APU 3017" in a
              for a in alertas), alertas


def test_el_cuadro_generado_alerta_el_pendiente_de_un_subapu(tmp_path):
    """El cuadro es el entregable: `generar_cuadro` congela la corrida antes de
    escribir el Excel, y la alerta tiene que sobrevivir esa vuelta por el snapshot."""
    import openpyxl
    alm = _alm_con_subapu(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_botadero=34))
    items = [LicitacionItem(item="1", descripcion="RELLENO", unidad="M3", cantidad=1,
                            precio_contractual=100000.0, shift="DIURNO")]
    cid = svc.construir_corrida(alm, "lic.xlsx", items, "DIURNO", False, carpeta_id=metro)
    svc.confirmar_item(alm, cid, 0, "4390", "DIURNO")
    out = svc.generar_cuadro(alm, cid)
    ws = openpyxl.load_workbook(out)["ALERTAS"]
    texto = "\n".join(str(c.value) for row in ws.iter_rows() for c in row
                      if c.value is not None)
    assert "distancia del proyecto no aplicada" in texto


def test_agregar_items_costea_con_la_distancia_del_proyecto(tmp_path):
    """Agregar una linea a una corrida ya armada tiene que costear con el MISMO
    contexto que el armado original (esta linea nueva == la linea equivalente que
    ya trae la corrida), no con la biblioteca cruda.

    Mientras el APU siga en la biblioteca, `_costear_row` relee su composicion
    fresca (con el contexto de la VISTA) para cualquier fila, asi que el costo
    vivo ya sale bien aunque el `Assembler` de `agregar_items` quedara sin
    `contexto=`. Donde SI se nota es en el respaldo que queda persistido con la
    linea (`row.componentes`, ver `_costear_row`): en cuanto el APU se borra de
    la biblioteca, el costeo cae a ese respaldo, y sin contexto trae la distancia
    CRUDA (26.25, -> 26250) en vez de la del proyecto (33.6, -> 33600) — costeando
    de menos, en silencio, justo cuando ya no hay de donde re-derivarla."""
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=32))
    cid = _corrida(alm, metro)                    # arma+confirma un RELLENO -> 33600
    original = svc.vista_corrida(alm, cid)["items"][0]
    assert original["costo_unitario"] == 33600     # confirma el punto de partida

    vista = svc.agregar_items(alm, cid, [LicitacionItem(
        item="2", descripcion="RELLENO", unidad="M3", cantidad=10,
        precio_contractual=100000.0, shift="DIURNO")])
    nueva = vista["items"][1]
    assert nueva["apu_codigo"] == "4390"           # el matcher la asigno sola
    assert nueva["costo_unitario"] == original["costo_unitario"] == 33600

    # El APU se borra de la biblioteca: la línea cae a su respaldo persistido.
    alm.apus.borrar_apu("4390", "DIURNO")
    tras_borrado = svc.vista_corrida(alm, cid)["items"][1]
    assert tras_borrado["costo_unitario"] == 33600  # el respaldo trajo la distancia del proyecto


def _alm_solo_peaje(base_dir):
    """Un APU cuyo ÚNICO componente es el peaje: un proyecto sin peaje lo deja SIN
    NADA, no solo más barato. Es el repro del bug crítico de la revisión."""
    alm = Almacen(precios_path=base_dir / "p.db", apus_path=base_dir / "a.db",
                  corridas_path=base_dir / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([
        Insumo(codigo="INT3", nombre="PEAJE", unidad="GLB", grupo="TRANSPORTES",
               precio=8000.0, fuente_precio="COSTO INTERNO")])
    alm.apus.insert_apus([Apu(codigo="9002", nombre="SOLO PEAJE", unidad="GLB",
                              shift="DIURNO", grupo="TRANSPORTES")])
    alm.apus.insert_components([
        ApuComponent(apu_codigo="9002", shift="DIURNO", insumo_codigo="INT3",
                     insumo_nombre="PEAJE", unidad="GLB", rendimiento=1.0,
                     precio_unitario_hist=8000.0)])
    return alm


def _corrida_peaje(alm, carpeta_id):
    items = [LicitacionItem(item="1", descripcion="SOLO PEAJE", unidad="GLB", cantidad=1,
                            precio_contractual=10000.0, shift="DIURNO")]
    cid = svc.construir_corrida(alm, "lic.xlsx", items, "DIURNO", False, carpeta_id=carpeta_id)
    svc.confirmar_item(alm, cid, 0, "9002", "DIURNO")
    return cid


def test_peaje_quitado_no_vuelve_por_el_respaldo_del_item(tmp_path):
    """Si el proyecto vacia la composicion, el costeo NO puede caer al respaldo de la
    composicion guardada: recobraria lo que el proyecto excluyo, sin alerta.

    Los dos ordenes de operacion tienen que dar el MISMO resultado: antes del fix,
    'corrida primero' costeaba 8000 sin alertas (el respaldo trajo de vuelta el
    peaje que el proyecto excluyo) y 'parametros primero' costeaba 0 con alerta --
    dos numeros distintos para el mismo estado."""
    # orden A: la corrida se arma ANTES de fijar los parámetros del proyecto.
    alm_a = _alm_solo_peaje(tmp_path / "orden_a")
    metro_a = alm_a.carpetas.crear("Metro")
    cid_a = _corrida_peaje(alm_a, metro_a)
    alm_a.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro_a, peaje_aplica=False))
    v_a = svc.vista_corrida(alm_a, cid_a)["items"][0]

    # orden B: los parámetros del proyecto se fijan ANTES de armar la corrida.
    alm_b = _alm_solo_peaje(tmp_path / "orden_b")
    metro_b = alm_b.carpetas.crear("Metro")
    alm_b.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro_b, peaje_aplica=False))
    cid_b = _corrida_peaje(alm_b, metro_b)
    v_b = svc.vista_corrida(alm_b, cid_b)["items"][0]

    for v in (v_a, v_b):
        assert v["costo_unitario"] == 0
        assert any("en $0" in a for a in v["alertas_costeo"]), v["alertas_costeo"]
    assert v_a["costo_unitario"] == v_b["costo_unitario"]
    assert v_a["alertas_costeo"] == v_b["alertas_costeo"]


def test_dos_quitar_que_vacian_un_apu_no_son_un_no_op(tmp_path):
    """Antes del fix, dos ajustes 'quitar' que vaciaban toda la composicion NO
    tenian efecto (el respaldo del item devolvia el costo de siempre)."""
    alm = Almacen(precios_path=tmp_path / "p.db", apus_path=tmp_path / "a.db",
                  corridas_path=tmp_path / "c.db")
    alm.init_schema()
    alm.precios.insert_insumos([
        Insumo(codigo="A1", nombre="INSUMO UNO", unidad="M3", grupo="X",
               precio=10000.0, fuente_precio="COSTO INTERNO"),
        Insumo(codigo="A2", nombre="INSUMO DOS", unidad="M3", grupo="X",
               precio=11900.0, fuente_precio="COSTO INTERNO")])
    alm.apus.insert_apus([Apu(codigo="8888", nombre="DOS INSUMOS", unidad="M3",
                              shift="DIURNO", grupo="X")])
    alm.apus.insert_components([
        ApuComponent(apu_codigo="8888", shift="DIURNO", insumo_codigo="A1",
                     insumo_nombre="INSUMO UNO", unidad="M3", rendimiento=1.0,
                     precio_unitario_hist=10000.0),
        ApuComponent(apu_codigo="8888", shift="DIURNO", insumo_codigo="A2",
                     insumo_nombre="INSUMO DOS", unidad="M3", rendimiento=1.0,
                     precio_unitario_hist=11900.0)])
    metro = alm.carpetas.crear("Metro")
    items = [LicitacionItem(item="1", descripcion="DOS INSUMOS", unidad="M3",
                            cantidad=1, precio_contractual=100000.0, shift="DIURNO")]
    cid = svc.construir_corrida(alm, "lic.xlsx", items, "DIURNO", False, carpeta_id=metro)
    svc.confirmar_item(alm, cid, 0, "8888", "DIURNO")
    assert svc.vista_corrida(alm, cid)["items"][0]["costo_unitario"] == 21900

    alm.carpetas.crear_ajuste(AjusteProyecto(
        apu_codigo="8888", shift="DIURNO", accion="quitar", insumo_codigo="A1",
        insumo_nombre="INSUMO UNO", carpeta_id=metro))
    alm.carpetas.crear_ajuste(AjusteProyecto(
        apu_codigo="8888", shift="DIURNO", accion="quitar", insumo_codigo="A2",
        insumo_nombre="INSUMO DOS", carpeta_id=metro))

    v = svc.vista_corrida(alm, cid)["items"][0]
    assert v["costo_unitario"] == 0
    assert any("en $0" in a for a in v["alertas_costeo"]), v["alertas_costeo"]


def test_el_cuadro_congelado_muestra_las_distancias_con_las_que_se_costeo(tmp_path):
    """Congelar a 34 km, cambiar el proyecto a 99, regenerar: el cuadro tiene que
    decir 34, que es con lo que estan calculados sus numeros."""
    import openpyxl
    alm = _alm(tmp_path)
    metro = alm.carpetas.crear("Metro")
    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=34))
    cid = _corrida(alm, metro)
    svc.congelar(alm, cid)
    congelado = svc.vista_corrida(alm, cid)["items"][0]["costo_unitario"]
    assert congelado == 35700                             # 1.05 * 34 * 1000

    alm.carpetas.set_parametros(ParametrosProyecto(carpeta_id=metro, km_granulares=99))

    # el encabezado de la vista sigue diciendo 34, no 99 (la corrida sigue congelada)
    vista = svc.vista_corrida(alm, cid)
    assert vista["transporte"]["km_granulares"] == 34
    assert vista["items"][0]["costo_unitario"] == congelado

    out = svc.generar_cuadro(alm, cid)
    wb = openpyxl.load_workbook(out)
    resumen = "\n".join(str(c.value) for row in wb["RESUMEN"].iter_rows() for c in row
                        if c.value is not None)
    assert "35700" in resumen or "35.700" in resumen       # RESUMEN sigue en 34km
    assert "103950" not in resumen and "103.950" not in resumen  # no el de 99km

    desv = "\n".join(str(c.value) for row in wb["DESVIACIONES DEL PROYECTO"].iter_rows()
                     for c in row if c.value is not None)
    assert "34" in desv and "99" not in desv               # y no se contradice a si mismo
